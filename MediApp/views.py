from datetime import datetime, timedelta
from decimal import Decimal
from pathlib import Path
import csv
import html
import json
import logging
import re
import urllib.error
import urllib.request

from django.core.mail import send_mail
from django.conf import settings
django_settings = settings
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q, Sum, Count
from django.db.models.functions import TruncDate
from django.utils import timezone
from django.urls import reverse
from django.views.decorators.http import require_POST 
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import Sale, SaleItem, Customer, Medicine, Batch

from .decorators import role_required
from .utils.email_utils import send_reminder_email
from .utils.data_loader import import_csv_dataset
from .forecasting import forecast_demand, get_all_medicine_forecasts, send_forecast_critical_stock_email

from .models import (
    User, Medicine, Batch, Customer, Supplier,
    Sale, SaleItem, MedicineReminder, InventoryLog, Category,
    Invoice, InvoiceItem, PurchaseOrder, PurchaseOrderItem,
    ExpiryReminderLog
)
from .forms import (
    UserRegistrationForm, MedicineForm, BatchForm,
    CustomerForm, SupplierForm, MedicineReminderForm, CategoryForm
)
from .decorators import role_required, admin_required, pharmacist_or_admin, assistant_or_above


def is_admin_role(user):
    """Return True for the global Admin role or a Django superuser."""
    return bool(getattr(user, 'is_superuser', False) or getattr(user, 'role', None) == 'admin')


def owner_scope_queryset(request, queryset, owner_field='created_by'):
    """Return a queryset that is admin-wide or user-private depending on the role."""
    if not getattr(request, 'user', None) or not request.user.is_authenticated:
        return queryset.none()
    if is_admin_role(request.user):
        return queryset
    return queryset.filter(**{owner_field: request.user})


def owner_scope_sales(request, queryset=None):
    """Scoped sale access for the current user."""
    queryset = queryset or Sale.objects.all()
    return owner_scope_queryset(request, queryset, 'created_by')


def owner_scope_purchase_orders(request, queryset=None):
    """Scoped purchase-order access for the current user."""
    queryset = queryset or PurchaseOrder.objects.all()
    return owner_scope_queryset(request, queryset, 'created_by')


# Authentication Views
def login_view(request):
    """Login view"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    error = None
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, f'Welcome back, {user.get_full_name()}!')
            return redirect('dashboard')
        else:
            error = 'Invalid username or password.'
    else:
        # Clear any stale 'Invalid username or password' flash messages from cookies
        storage = messages.get_messages(request)
        other_messages = [m for m in storage if 'Invalid username or password' not in str(m)]
        for m in other_messages:
            messages.add_message(request, m.level, m.message, extra_tags=m.extra_tags)
    
    return render(request, 'login.html', {'error': error})


def logout_view(request):
    """Logout view"""
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('login')


@admin_required
def register_view(request):
    """Register new user (Admin only)"""
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f'✅ User "{user.username}" ({user.get_role_display()}) created successfully!')
            return redirect('register')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = UserRegistrationForm()

    return render(request, 'register.html', {'form': form})


def _ensure_all_sales_have_invoices():
    """Ensure every completed sale in the database has an Invoice record."""
    missing_sales = Sale.objects.filter(invoice__isnull=True).order_by('date', 'id').select_related('customer', 'created_by').prefetch_related('items__medicine', 'items__batch')
    if not missing_sales.exists():
        return

    pharmacy_name = getattr(django_settings, 'PHARMACY_NAME', 'PharmaCare Pharmacy')
    pharmacy_address = getattr(django_settings, 'PHARMACY_ADDRESS', 'Pharmacy address')
    pharmacy_gstin = getattr(django_settings, 'PHARMACY_GSTIN', '')
    pharmacy_phone = getattr(django_settings, 'PHARMACY_PHONE', '')

    default_user = User.objects.filter(is_superuser=True).first() or User.objects.first()

    with transaction.atomic():
        for s in missing_sales:
            year = s.date.year if s.date else timezone.now().year
            inv_num = f"INV-{year}-{s.id:05d}"

            items = list(s.items.all())
            subtotal_without_gst = Decimal('0')
            gst_total = Decimal('0')

            for item in items:
                if item.price is not None and item.quantity is not None:
                    gross = Decimal(item.price) * Decimal(item.quantity)
                    base = gross / Decimal('1.05')
                    subtotal_without_gst += base
                    gst_total += (gross - base)

            cgst = gst_total / Decimal('2')
            sgst = gst_total / Decimal('2')

            customer = s.customer
            inv = Invoice.objects.create(
                invoice_number=inv_num,
                sale=s,
                customer=customer,
                created_by=s.created_by or default_user,
                invoice_date=s.date,
                pharmacy_name=pharmacy_name,
                pharmacy_address=pharmacy_address,
                pharmacy_gstin=pharmacy_gstin,
                pharmacy_phone=pharmacy_phone,
                customer_name=customer.name if customer else 'Walk-in Customer',
                customer_phone=customer.contact_number if customer else '',
                payment_method=s.payment_method or 'Cash',
                subtotal=subtotal_without_gst,
                gst_total=gst_total,
                cgst=cgst,
                sgst=sgst,
                discount=s.discount or Decimal('0'),
                final_amount=s.total_price or Decimal('0'),
            )

            for it in items:
                gross = Decimal(it.price) * Decimal(it.quantity)
                base = gross / Decimal('1.05')
                gst = gross - base
                InvoiceItem.objects.create(
                    invoice=inv,
                    medicine=it.medicine,
                    batch=it.batch,
                    medicine_name=it.medicine.name if it.medicine else 'Unknown Medicine',
                    batch_number=it.batch.batch_name if it.batch else '',
                    expiry_date=it.batch.expiry_date if it.batch else None,
                    quantity=it.quantity,
                    price=it.price,
                    gst_percent=Decimal('5'),
                    gst_amount=gst,
                    line_total=gross,
                )


@assistant_or_above
def invoice_history(request):
    """Invoice history/listing with real-time statistics, search, and server-side pagination."""
    # Ensure every sale in database has an invoice
    _ensure_all_sales_have_invoices()

    query = request.GET.get('q', '').strip()
    filter_type = request.GET.get('filter', 'all')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    today = timezone.now().date()

    invoices_qs = owner_scope_queryset(
        request,
        Invoice.objects.select_related('sale', 'customer', 'created_by').prefetch_related('invoice_items'),
        'created_by'
    )

    # Date filter presets
    if filter_type == 'today':
        invoices_qs = invoices_qs.filter(invoice_date__date=today)
    elif filter_type == 'week':
        week_ago = today - timedelta(days=7)
        invoices_qs = invoices_qs.filter(invoice_date__date__gte=week_ago, invoice_date__date__lte=today)
    elif filter_type == 'month':
        invoices_qs = invoices_qs.filter(invoice_date__year=today.year, invoice_date__month=today.month)
    elif filter_type == 'year':
        invoices_qs = invoices_qs.filter(invoice_date__year=today.year)
    elif filter_type == 'custom':
        if start_date and end_date:
            invoices_qs = invoices_qs.filter(invoice_date__date__gte=start_date, invoice_date__date__lte=end_date)
        elif start_date:
            invoices_qs = invoices_qs.filter(invoice_date__date__gte=start_date)
        elif end_date:
            invoices_qs = invoices_qs.filter(invoice_date__date__lte=end_date)
    else:
        if start_date:
            invoices_qs = invoices_qs.filter(invoice_date__date__gte=start_date)
        if end_date:
            invoices_qs = invoices_qs.filter(invoice_date__date__lte=end_date)

    if query:
        invoices_qs = invoices_qs.filter(
            Q(invoice_number__icontains=query) |
            Q(customer_name__icontains=query) |
            Q(customer_phone__icontains=query) |
            Q(invoice_items__medicine_name__icontains=query) |
            Q(invoice_items__batch_number__icontains=query)
        ).distinct()

    # Calculate summary metrics
    total_invoices_count = invoices_qs.count()
    all_time_invoices_count = owner_scope_queryset(request, Invoice.objects.all(), 'created_by').count()
    today_invoices_count = owner_scope_queryset(request, Invoice.objects.filter(invoice_date__date=today), 'created_by').count()

    total_billed_amount = invoices_qs.aggregate(Sum('final_amount'))['final_amount__sum'] or Decimal('0.00')
    total_gst_amount = invoices_qs.aggregate(Sum('gst_total'))['gst_total__sum'] or Decimal('0.00')

    # Server-side pagination (25 records per page for ultra-fast loading)
    page = request.GET.get('page', 1)
    paginator = Paginator(invoices_qs, 25)
    try:
        invoices = paginator.page(page)
    except PageNotAnInteger:
        invoices = paginator.page(1)
    except EmptyPage:
        invoices = paginator.page(paginator.num_pages)

    return render(request, 'billing/invoice_history.html', {
        'invoices': invoices,
        'page_obj': invoices,
        'paginator': paginator,
        'is_paginated': invoices.has_other_pages(),
        'total_invoices_count': total_invoices_count,
        'all_time_invoices_count': all_time_invoices_count,
        'today_invoices_count': today_invoices_count,
        'total_billed_amount': total_billed_amount,
        'total_gst_amount': total_gst_amount,
        'search_query': query,
        'filter_type': filter_type,
        'start_date': start_date,
        'end_date': end_date,
    })


@assistant_or_above
def invoice_detail(request, invoice_id):
    """View and print a single invoice."""
    invoice = get_object_or_404(
        owner_scope_queryset(
            request,
            Invoice.objects.select_related('sale', 'customer', 'created_by').prefetch_related('invoice_items__medicine', 'invoice_items__batch'),
            'created_by'
        ),
        id=invoice_id,
    )
    return render(request, 'billing/invoice_detail.html', {'invoice': invoice})


@assistant_or_above
def invoice_print(request, invoice_id):
    """A print-friendly invoice view."""
    invoice = get_object_or_404(
        owner_scope_queryset(
            request,
            Invoice.objects.select_related('sale', 'customer', 'created_by').prefetch_related('invoice_items__medicine', 'invoice_items__batch'),
            'created_by'
        ),
        id=invoice_id,
    )
    return render(request, 'billing/invoice_print.html', {'invoice': invoice})


@assistant_or_above
def purchase_order_list(request):
    """List all owner-scoped purchase orders."""
    purchase_orders = owner_scope_purchase_orders(
        request,
        PurchaseOrder.objects.select_related('supplier', 'created_by').prefetch_related('items__medicine')
    )
    return render(request, 'purchase_orders/purchase_order_list.html', {
        'purchase_orders': purchase_orders,
    })


@pharmacist_or_admin
def purchase_order_add(request):
    """Create a purchase order for one or more medicines."""
    suppliers = owner_scope_queryset(request, Supplier.objects.all(), 'created_by')
    medicines = owner_scope_queryset(request, Medicine.objects.select_related('supplier', 'category').prefetch_related('batches'), 'created_by')

    if request.method == 'POST':
        supplier_id = request.POST.get('supplier')
        supplier = get_object_or_404(owner_scope_queryset(request, Supplier.objects.all(), 'created_by'), id=supplier_id)
        medicine_ids = request.POST.getlist('medicine_id[]')
        quantities = request.POST.getlist('quantity[]')
        cost_prices = request.POST.getlist('cost_price[]')

        order_number = _next_purchase_order_number()
        po = PurchaseOrder.objects.create(
            order_number=order_number,
            supplier=supplier,
            status='Pending',
            order_date=timezone.now(),
            expected_date=request.POST.get('expected_date') or None,
            created_by=request.user,
            total_amount=Decimal('0')
        )

        total_amount = Decimal('0')
        for idx, medicine_id in enumerate(medicine_ids):
            if not medicine_id:
                continue
            medicine = get_object_or_404(owner_scope_queryset(request, Medicine.objects.all(), 'created_by'), id=medicine_id)
            quantity = int(quantities[idx] or 0)
            if quantity <= 0:
                continue
            cost_price = Decimal(cost_prices[idx] or '0')
            PurchaseOrderItem.objects.create(
                purchase_order=po,
                medicine=medicine,
                quantity=quantity,
                cost_price=cost_price
            )
            total_amount += cost_price * Decimal(quantity)

        po.total_amount = total_amount
        po.save(update_fields=['total_amount'])
        messages.success(request, f'Purchase order {po.order_number} created successfully!')
        return redirect('purchase_order_list')

    return render(request, 'purchase_orders/purchase_order_add.html', {
        'suppliers': suppliers,
        'medicines': medicines,
    })


@pharmacist_or_admin
@require_POST
def purchase_order_receive(request, po_id):
    """Receive a purchase order and add stock to the relevant medicine batches."""
    po = get_object_or_404(owner_scope_purchase_orders(request, PurchaseOrder.objects.select_related('supplier')), id=po_id)
    if po.status == 'Received':
        messages.info(request, f'Purchase order {po.order_number} is already received.')
        return redirect('purchase_order_list')

    today = timezone.now().date()
    for item in po.items.select_related('medicine'):
        if not item.medicine:
            continue
        Batch.objects.create(
            medicine=item.medicine,
            batch_name=f"PO-{po.order_number}",
            add_date=today,
            expiry_date=today + timedelta(days=365),
            quantity=item.quantity,
            purchase_price=item.cost_price,
            created_by=request.user,
        )
        InventoryLog.objects.create(
            medicine=item.medicine,
            batch=item.medicine.batches.order_by('-id').first(),
            action='add',
            quantity_change=item.quantity,
            performed_by=request.user,
            notes=f'Received against purchase order {po.order_number}'
        )

    po.status = 'Received'
    po.received_date = today
    po.save(update_fields=['status', 'received_date'])
    messages.success(request, f'Purchase order {po.order_number} marked as received.')
    return redirect('purchase_order_list')


def _next_purchase_order_number():
    """Return a new purchase order number like PO-2026-0001."""
    year = timezone.now().year
    prefix = f'PO-{year}-'
    last_po = PurchaseOrder.objects.filter(order_number__startswith=prefix).order_by('-order_number').first()
    if last_po:
        try:
            last_sequence = int(last_po.order_number.split('-')[-1])
        except (ValueError, IndexError):
            last_sequence = 0
    else:
        last_sequence = 0
    return f'{prefix}{last_sequence + 1:04d}'


# Dashboard View
@assistant_or_above
def dashboard(request):
    """Main dashboard view"""
    today = timezone.now().date()
    duration = request.GET.get('duration', 'today')
    
    start_date = today
    if duration == 'week':
        start_date = today - timedelta(days=7)
    elif duration == 'month':
        start_date = today - timedelta(days=30)
    elif duration == 'year':
        start_date = today - timedelta(days=365)
    
    # Auto-import CSV dataset when the app is running against an empty dataset.
    if not (Category.objects.exists() or Supplier.objects.exists() or Medicine.objects.exists() or Customer.objects.exists() or Sale.objects.exists()):
        import_csv_dataset()

    # Global statistics, scoped to the signed-in user unless admin sees everyone
    scoped_medicines = owner_scope_queryset(request, Medicine.objects.all(), 'created_by')
    scoped_customers = owner_scope_queryset(request, Customer.objects.all(), 'created_by')
    scoped_suppliers = owner_scope_queryset(request, Supplier.objects.all(), 'created_by')
    scoped_categories = owner_scope_queryset(request, Category.objects.all(), 'created_by')
    scoped_sales = owner_scope_sales(request, Sale.objects.all())
    scoped_sale_items = SaleItem.objects.filter(sale__in=scoped_sales)
    scoped_batches = owner_scope_queryset(request, Batch.objects.all(), 'created_by')

    total_medicines = scoped_medicines.count()
    total_customers = scoped_customers.count()
    total_suppliers = scoped_suppliers.count()
    total_categories = scoped_categories.count()
    total_sales = scoped_sales.count()
    all_time_revenue = scoped_sales.aggregate(Sum('total_price'))['total_price__sum'] or 0
    low_stock_count = sum(1 for med in scoped_medicines.all() if med.is_low_stock)
    
    # Period-specific statistics
    period_sales = scoped_sales.filter(date__date__gte=start_date)
    period_sales_count = period_sales.count()
    period_revenue = period_sales.aggregate(Sum('total_price'))['total_price__sum'] or 0
    period_new_medicines = scoped_medicines.filter(created_at__date__gte=start_date).count()
    period_new_suppliers = scoped_suppliers.filter(created_at__date__gte=start_date).count()
    
    # Expiry statistics & dynamic labels based on duration filter
    if duration == 'week':
        expired_batches = scoped_batches.filter(
            expiry_date__lte=today,
            expiry_date__gte=today - timedelta(days=7),
            quantity__gt=0
        )
        expiring_soon = scoped_batches.filter(
            expiry_date__gte=today,
            expiry_date__lte=today + timedelta(days=7),
            quantity__gt=0
        )
        expiring_label = 'Expiring in 7 Days'
        expiring_subtext = 'Batches expiring in 7 days'
        expired_subtext = 'Expired in last 7 days'
    elif duration == 'month':
        expired_batches = scoped_batches.filter(
            expiry_date__lte=today,
            expiry_date__gte=today - timedelta(days=30),
            quantity__gt=0
        )
        expiring_soon = scoped_batches.filter(
            expiry_date__gte=today,
            expiry_date__lte=today + timedelta(days=30),
            quantity__gt=0
        )
        expiring_label = 'Expiring in 30 Days'
        expiring_subtext = 'Batches expiring in 30 days'
        expired_subtext = 'Expired in last 30 days'
    elif duration == 'year':
        expired_batches = scoped_batches.filter(
            expiry_date__lte=today,
            expiry_date__gte=today - timedelta(days=365),
            quantity__gt=0
        )
        expiring_soon = scoped_batches.filter(
            expiry_date__gte=today,
            expiry_date__lte=today + timedelta(days=365),
            quantity__gt=0
        )
        expiring_label = 'Expiring This Year'
        expiring_subtext = 'Batches expiring this year'
        expired_subtext = 'Expired this year'
    else:  # today
        expired_batches = scoped_batches.filter(
            expiry_date__lte=today,
            expiry_date__gte=today,
            quantity__gt=0
        )
        expiring_soon = scoped_batches.filter(
            expiry_date=today,
            quantity__gt=0
        )
        expiring_label = 'Expiring Today'
        expiring_subtext = 'Batches expiring today'
        expired_subtext = 'Expired today'
    
    # Recent items
    recent_logs = InventoryLog.objects.filter(
        performed_by=request.user
    ).select_related('medicine', 'batch', 'performed_by')[:10]
    recent_sales = scoped_sales.select_related('customer', 'created_by').prefetch_related('items')[:10]
    low_stock_medicines = [med for med in scoped_medicines.all() if med.is_low_stock][:10]
    
    context = {
        'duration': duration,
        'total_medicines': total_medicines,
        'total_customers': total_customers,
        'total_suppliers': total_suppliers,
        'total_categories': total_categories,
        'total_sales': total_sales,
        'all_time_revenue': all_time_revenue,
        'low_stock_count': low_stock_count,
        'period_sales_count': period_sales_count,
        'period_revenue': period_revenue,
        'period_new_medicines': period_new_medicines,
        'period_new_suppliers': period_new_suppliers,
        'expired_count': expired_batches.count(),
        'expiring_soon_count': expiring_soon.count(),
        'expiring_label': expiring_label,
        'expiring_subtext': expiring_subtext,
        'expired_subtext': expired_subtext,
        'recent_logs': recent_logs,
        'recent_sales': recent_sales,
        'low_stock_medicines': low_stock_medicines,
    }
    
    return render(request, 'dashboard.html', context)


# Medicine Views
@pharmacist_or_admin
def medicine_list(request):
    """List all medicines"""
    search_query = request.GET.get('search', '')
    expiry_filter = request.GET.get('expiry', '')
    today = timezone.now().date()
    medicines = owner_scope_queryset(request, Medicine.objects.select_related('supplier').prefetch_related('batches'), 'created_by')

    if search_query:
        medicines = medicines.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(supplier__name__icontains=search_query)
        )

    if expiry_filter in ['7', '15', '30']:
        days = int(expiry_filter)
        target_date = today + timedelta(days=days)
        # Filter medicines that have any batch expiring within the chosen window
        medicines = medicines.filter(
            batches__expiry_date__gte=today,
            batches__expiry_date__lte=target_date
        ).distinct()
        # Attach the batch that falls within THIS specific window to each medicine
        # so the card shows the relevant batch, not just the globally nearest one.
        medicines = list(medicines)
        for med in medicines:
            med.display_batch = (
                med.batches
                .filter(expiry_date__gte=today, expiry_date__lte=target_date, quantity__gt=0)
                .order_by('expiry_date')
                .first()
            )
    else:
        # No expiry filter — attach the nearest active batch for each medicine
        medicines = list(medicines)
        for med in medicines:
            med.display_batch = med.nearest_active_batch

    context = {
        'medicines': medicines,
        'search_query': search_query,
        'expiry_filter': expiry_filter,
    }
    return render(request, 'medicine/medicine_list.html', context)


@pharmacist_or_admin
def export_all_medicine_stock(request):
    """Export all medicine stock to Excel/CSV"""
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="supplied_medicine_stock.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Medicine Name', 'Description', 'Category', 'Supplier', 'Supplier Contact', 
        'Supplier Email', 'Price (Selling)', 'Batch Name', 'Purchase Price', 
        'Quantity in Stock', 'Expiry Date', 'Is Expired', 'Total Medicine Stock', 
        'Low Stock Warning'
    ])
    
    medicines = owner_scope_queryset(request, Medicine.objects.all().prefetch_related('batches', 'supplier', 'category'), 'created_by')
    for med in medicines:
        supplier_name = med.supplier.name if med.supplier else 'No Supplier'
        supplier_contact = med.supplier.contact_number if med.supplier else 'N/A'
        supplier_email = med.supplier.email if med.supplier else 'N/A'
        category_name = med.category.name if med.category else 'Uncategorized'
        
        batches = med.batches.all()
        if not batches:
            writer.writerow([
                med.name, med.description, category_name, supplier_name, supplier_contact,
                supplier_email, float(med.price), 'N/A', 'N/A', 0, 'N/A', 'N/A', 0, 'Yes'
            ])
            continue
            
        for batch in batches:
            writer.writerow([
                med.name, med.description, category_name, supplier_name, supplier_contact,
                supplier_email, float(med.price), batch.batch_name, 
                float(batch.purchase_price) if batch.purchase_price else 0,
                batch.quantity, 
                batch.expiry_date.strftime('%Y-%m-%d') if batch.expiry_date else 'N/A',
                'Yes' if batch.is_expired else 'No',
                med.total_quantity,
                'Yes' if med.is_low_stock else 'No'
            ])
            
    return response


@pharmacist_or_admin
def add_medicine(request):
    """Add new medicine with batches"""
    if request.method == 'POST':
        medicine_form = MedicineForm(request.POST)

        if medicine_form.is_valid():
            medicine = medicine_form.save(commit=False)
            medicine.created_by = request.user
            medicine.save()
            medicine_form.save_m2m()

            # Handle batch data from POST
            batch_names = request.POST.getlist('batch_name[]')
            add_dates = request.POST.getlist('add_date[]')
            expiry_dates = request.POST.getlist('expiry_date[]')
            quantities = request.POST.getlist('quantity[]')
            purchase_prices = request.POST.getlist('purchase_price[]')

            for i in range(len(batch_names)):
                if batch_names[i]:
                    Batch.objects.create(
                        medicine=medicine,
                        batch_name=batch_names[i],
                        add_date=add_dates[i],
                        expiry_date=expiry_dates[i],
                        quantity=quantities[i],
                        purchase_price=purchase_prices[i] if i < len(purchase_prices) else 0,
                        created_by=request.user
                    )

                    # Create inventory log
                    InventoryLog.objects.create(
                        medicine=medicine,
                        batch=medicine.batches.last(),
                        action='add',
                        quantity_change=int(quantities[i]),
                        performed_by=request.user,
                        notes=f'Initial batch {batch_names[i]} added'
                    )

            messages.success(request, f'Medicine "{medicine.name}" added successfully!')
            return redirect('medicine_list')
    else:
        medicine_form = MedicineForm()

    suppliers = owner_scope_queryset(request, Supplier.objects.all(), 'created_by')
    return render(request, 'medicine/add_medicine.html', {
        'medicine_form': medicine_form,
        'suppliers': suppliers
    })


@pharmacist_or_admin
def edit_medicine(request, id):
    """Edit medicine details"""
    medicine = get_object_or_404(Medicine, id=id)
    
    if request.method == 'POST':
        form = MedicineForm(request.POST, instance=medicine)
        if form.is_valid():
            form.save()
            messages.success(request, f'Medicine "{medicine.name}" updated successfully!')
            return redirect('medicine_list')
    else:
        form = MedicineForm(instance=medicine)
    
    return render(request, 'medicine/edit_medicine.html', {
        'form': form,
        'medicine': medicine
    })


@pharmacist_or_admin
def delete_medicine(request, id):
    """Delete medicine"""
    medicine = get_object_or_404(Medicine, id=id)
    name = medicine.name
    medicine.delete()
    messages.success(request, f'Medicine "{name}" deleted successfully!')
    return redirect('medicine_list')


@pharmacist_or_admin
def get_batch(request, id):
    batch = get_object_or_404(Batch, id=id)
    return JsonResponse({
        'id': batch.id,
        'batch_name': batch.batch_name,
        'add_date': batch.add_date.strftime('%Y-%m-%d'),
        'expiry_date': batch.expiry_date.strftime('%Y-%m-%d'),
        'quantity': batch.quantity,
        'purchase_price': float(batch.purchase_price),
    })


@pharmacist_or_admin
def save_batch(request):
    if request.method == 'POST':
        batch_id = request.POST.get('batch_id')
        medicine_id = request.POST.get('medicine_id')
        batch_name = request.POST.get('batch_name')
        add_date = request.POST.get('add_date')
        expiry_date = request.POST.get('expiry_date')
        quantity = request.POST.get('quantity')
        purchase_price = request.POST.get('purchase_price', 0)

        medicine = get_object_or_404(Medicine, id=medicine_id)

        if batch_id:
            batch = get_object_or_404(Batch, id=batch_id)
            old_quantity = batch.quantity
            batch.batch_name = batch_name
            batch.add_date = add_date
            batch.expiry_date = expiry_date
            batch.quantity = quantity
            batch.purchase_price = purchase_price
            batch.save()
            
            # Log the update
            quantity_diff = int(quantity) - old_quantity
            if quantity_diff != 0:
                InventoryLog.objects.create(
                    medicine=medicine,
                    batch=batch,
                    action='update',
                    quantity_change=quantity_diff,
                    performed_by=request.user,
                    notes=f'Batch {batch_name} quantity updated'
                )
        else:
            batch = Batch.objects.create(
                medicine=medicine,
                batch_name=batch_name,
                add_date=add_date,
                expiry_date=expiry_date,
                quantity=quantity,
                purchase_price=purchase_price
            )
            
            # Log the addition
            InventoryLog.objects.create(
                medicine=medicine,
                batch=batch,
                action='add',
                quantity_change=int(quantity),
                performed_by=request.user,
                notes=f'New batch {batch_name} added'
            )

        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': 'Invalid request'})


@pharmacist_or_admin
def delete_batch(request, id):
    if request.method == 'POST':
        batch = get_object_or_404(Batch, id=id)
        batch.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': 'Invalid request'})


# Sales Views
@assistant_or_above
def sales_list(request):
    """List all sales with filtering and fast server-side pagination"""
    sales_qs = owner_scope_sales(request, Sale.objects.select_related('customer', 'created_by', 'invoice').prefetch_related('items__medicine', 'items__batch'))
    
    today = timezone.now().date()
    
    # Filter by date range
    filter_type = request.GET.get('filter', 'all')
    if filter_type == 'today':
        sales_qs = sales_qs.filter(date__date=today)
    elif filter_type == 'month':
        sales_qs = sales_qs.filter(date__month=today.month, date__year=today.year)
    elif filter_type == 'custom':
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date and end_date:
            sales_qs = sales_qs.filter(date__date__range=[start_date, end_date])
    
    # Search
    search_query = request.GET.get('search', '').strip()
    if search_query:
        sales_qs = sales_qs.filter(
            Q(id__icontains=search_query) |
            Q(customer__name__icontains=search_query) |
            Q(customer__contact_number__icontains=search_query)
        )
    
    # Calculate totals
    total_sales_count = sales_qs.count()
    today_sales_count = owner_scope_sales(request, Sale.objects.filter(date__date=today)).count()

    # Calculate total discount given
    total_discount = sales_qs.aggregate(Sum('discount'))['discount__sum'] or 0

    # Paginate (25 records per page for lightning fast load time)
    page = request.GET.get('page', 1)
    paginator = Paginator(sales_qs, 25)
    try:
        sales = paginator.page(page)
    except PageNotAnInteger:
        sales = paginator.page(1)
    except EmptyPage:
        sales = paginator.page(paginator.num_pages)

    context = {
        'sales': sales,
        'paginator': paginator,
        'page_obj': sales,
        'is_paginated': sales.has_other_pages(),
        'total_sales_count': total_sales_count,
        'today_sales_count': today_sales_count,
        'total_discount': total_discount,
        'filter_type': filter_type,
        'search_query': search_query,
    }
    return render(request, 'sales/sales_list.html', context)



def _next_invoice_number():
    """Return a new invoice number such as INV-2026-0001."""
    year = timezone.now().year
    prefix = f'INV-{year}-'
    last_invoice = Invoice.objects.filter(invoice_number__startswith=prefix).order_by('-invoice_number').first()
    if last_invoice:
        try:
            last_sequence = int(last_invoice.invoice_number.split('-')[-1])
        except (ValueError, IndexError):
            last_sequence = 0
    else:
        last_sequence = 0
    return f'{prefix}{last_sequence + 1:04d}'


def generate_invoice_for_sale(request, sale):
    """Persist a GST invoice payload for a completed sale."""
    if Invoice.objects.filter(sale=sale).exists():
        return Invoice.objects.get(sale=sale)

    items = list(sale.items.select_related('medicine', 'batch'))
    subtotal_without_gst = Decimal('0')
    gst_total = Decimal('0')
    cgst = Decimal('0')
    sgst = Decimal('0')

    for item in items:
        if item.price is None or item.quantity is None:
            continue
        # The UI stores an MRP that already includes GST. Split it back to base + GST.
        gst_rate = Decimal('0.05')
        gross_line_value = Decimal(item.price) * Decimal(item.quantity)
        base_value = gross_line_value / (Decimal('1') + gst_rate)
        line_gst = gross_line_value - base_value
        subtotal_without_gst += base_value
        gst_total += line_gst

    cgst = gst_total / Decimal('2')
    sgst = gst_total / Decimal('2')

    customer = sale.customer
    invoice = Invoice.objects.create(
        invoice_number=_next_invoice_number(),
        sale=sale,
        customer=customer,
        created_by=request.user,
        invoice_date=sale.date,
        pharmacy_name=getattr(settings, 'PHARMACY_NAME', 'PharmaCare Pharmacy'),
        pharmacy_address=getattr(settings, 'PHARMACY_ADDRESS', 'Pharmacy address'),
        pharmacy_gstin=getattr(settings, 'PHARMACY_GSTIN', ''),
        pharmacy_phone=getattr(settings, 'PHARMACY_PHONE', ''),
        customer_name=customer.name if customer else 'Walk-in Customer',
        customer_phone=customer.contact_number if customer else '',
        payment_method=sale.payment_method,
        subtotal=subtotal_without_gst,
        gst_total=gst_total,
        cgst=cgst,
        sgst=sgst,
        discount=sale.discount,
        final_amount=sale.total_price,
    )

    for item in items:
        if item.price is None or item.quantity is None:
            continue
        gross_line_value = Decimal(item.price) * Decimal(item.quantity)
        base_value = gross_line_value / Decimal('1.05')
        line_gst = gross_line_value - base_value
        InvoiceItem.objects.create(
            invoice=invoice,
            medicine=item.medicine,
            batch=item.batch,
            medicine_name=item.medicine.name if item.medicine else 'Unknown Medicine',
            batch_number=item.batch.batch_name if item.batch else '',
            expiry_date=item.batch.expiry_date if item.batch else None,
            quantity=item.quantity,
            price=item.price,
            gst_percent=Decimal('5.00'),
            gst_amount=line_gst,
            line_total=gross_line_value,
        )

    return invoice


@assistant_or_above
def add_sale(request):
    """Add new sale with discount support"""
    if request.method == 'POST':
        customer_id = request.POST.get('customer_id')
        customer = None
        
        if customer_id and customer_id != 'walk-in':
            customer = get_object_or_404(Customer, id=customer_id)
        
        # Get discount value (default to 0 if not provided or invalid)
        try:
            discount = Decimal(request.POST.get('discount', '0') or '0')
            if discount < 0:
                discount = Decimal('0')
        except:
            discount = Decimal('0')
        
        # Get payment method
        payment_method = request.POST.get('payment_method', 'Cash')

        # Create sale with discount
        sale = Sale.objects.create(
            customer=customer,
            created_by=request.user,
            total_price=0,
            discount=discount,
            payment_method=payment_method,
            status='Completed'
        )
        
        # Add sale items
        medicine_ids = request.POST.getlist('medicine_id[]')
        batch_ids = request.POST.getlist('batch_id[]')
        quantities = request.POST.getlist('quantity[]')
        prices = request.POST.getlist('price[]')
        
        total = Decimal('0')
        for i in range(len(medicine_ids)):
            if medicine_ids[i]:
                medicine = get_object_or_404(Medicine, id=medicine_ids[i])
                batch = get_object_or_404(Batch, id=batch_ids[i])
                quantity = int(quantities[i])
                price = Decimal(prices[i])
                
                # Create sale item
                SaleItem.objects.create(
                    sale=sale,
                    medicine=medicine,
                    batch=batch,
                    quantity=quantity,
                    price=price,
                    cost_price=batch.purchase_price
                )
                
                # Update batch quantity
                batch.quantity -= quantity
                batch.save()
                
                # Create inventory log
                InventoryLog.objects.create(
                    medicine=medicine,
                    batch=batch,
                    action='sale',
                    quantity_change=-quantity,
                    performed_by=request.user,
                    notes=f'Sold in sale #{sale.id}'
                )
                
                total += price * quantity
        
        # Update sale total (subtotal - discount)
        sale.total_price = total - discount
        sale.save()

        # Create a GST-aligned invoice record for this sale
        invoice = generate_invoice_for_sale(request, sale)

        messages.success(request, f'Sale #{sale.id} created successfully! Invoice {invoice.invoice_number} is ready.')
        return redirect('invoice_detail', invoice_id=invoice.id)
    
    medicines = owner_scope_queryset(request, Medicine.objects.all(), 'created_by')
    customers = owner_scope_queryset(request, Customer.objects.all(), 'created_by')
    return render(request, 'sales/add_sale.html', {
        'medicines': medicines,
        'customers': customers
    })


@assistant_or_above
def delete_sale(request, id):
    """Delete sale"""
    sale = get_object_or_404(owner_scope_sales(request, Sale.objects.all()), id=id)
    sale.delete()
    messages.success(request, f'Sale #{id} deleted successfully!')
    return redirect('sales_list')


@assistant_or_above
@require_POST
def refund_sale(request, id):
    """Refund a sale and restore inventory"""
    sale = get_object_or_404(owner_scope_sales(request, Sale.objects.all()), id=id)
    
    if sale.status == 'Refunded':
        messages.warning(request, f'Sale #{id} is already refunded.')
        return redirect('sales_list')
        
    # Process refund for each item
    for item in sale.items.all():
        if item.batch:
            item.batch.quantity += item.quantity
            item.batch.save()
            
            InventoryLog.objects.create(
                medicine=item.medicine,
                batch=item.batch,
                action='refund',
                quantity_change=item.quantity,
                performed_by=request.user,
                notes=f'Refunded from sale #{sale.id}'
            )
            
    sale.status = 'Refunded'
    sale.save()
    
    messages.success(request, f'Sale #{id} has been successfully refunded and inventory restored.')
    return redirect('sales_list')
    return redirect('sales_list')


@assistant_or_above
def customer_purchase_history_json(request, customer_id):
    """Return the 5 most recent purchases for a customer as JSON (for the Add Sale preview panel)."""
    customer = get_object_or_404(owner_scope_queryset(request, Customer.objects.all(), 'created_by'), id=customer_id)
    sales = (
        owner_scope_sales(request, customer.sales.all())
        .prefetch_related('items__medicine')
        .order_by('-date')[:5]
    )
    data = []
    for sale in sales:
        data.append({
            'id': sale.id,
            'date': sale.date.strftime('%d %b %Y'),
            'total': str(sale.total_price),
            'status': sale.status,
            'payment_method': sale.payment_method,
            'items': [
                {
                    'name': item.medicine.name if item.medicine else '(removed)',
                    'quantity': item.quantity,
                    'price': str(item.price),
                }
                for item in sale.items.all()
            ],
        })
    return JsonResponse({'customer_name': customer.name, 'sales': data})


# Customer Views
@assistant_or_above
def customer_list(request):
    """List all customers with fast search and pagination"""
    search_query = request.GET.get('search', '').strip()
    customers_qs = owner_scope_queryset(request, Customer.objects.prefetch_related('reminders'), 'created_by')
    
    if search_query:
        customers_qs = customers_qs.filter(
            Q(name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(contact_number__icontains=search_query)
        )
    
    total_count = customers_qs.count()
    page = request.GET.get('page', 1)
    paginator = Paginator(customers_qs, 24)
    try:
        customers = paginator.page(page)
    except PageNotAnInteger:
        customers = paginator.page(1)
    except EmptyPage:
        customers = paginator.page(paginator.num_pages)
    
    context = {
        'customers': customers,
        'paginator': paginator,
        'page_obj': customers,
        'is_paginated': customers.has_other_pages(),
        'total_count': total_count,
        'search_query': search_query,
    }
    return render(request, 'customer/customer_list.html', context)



@assistant_or_above
def add_customer(request):
    """Add new customer """
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer = form.save(commit=False)
            customer.created_by = request.user
            customer.save()
            messages.success(request, f'Customer "{customer.name}" added successfully!')
            return redirect('customer_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CustomerForm()

    return render(request, 'customer/add_customer.html', {'form': form})


@assistant_or_above
def edit_customer(request, customer_id):
    customer = get_object_or_404(owner_scope_queryset(request, Customer.objects.all(), 'created_by'), id=customer_id)

    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, 'Customer updated.')
            return redirect('customer_list')
    else:
        form = CustomerForm(instance=customer)

    # pass medicines to select list in modal
    medicines = owner_scope_queryset(request, Medicine.objects.order_by('name').all(), 'created_by')

    # Purchase history: all sales linked to this customer, newest first
    purchase_history = (
        customer.sales
        .prefetch_related('items__medicine', 'items__batch')
        .select_related('created_by')
        .order_by('-date')
    )

    total_spent = sum(s.total_price for s in purchase_history if s.status == 'Completed')
    total_visits = purchase_history.filter(status='Completed').count()

    return render(request, 'customer/edit_customer.html', {
        'customer': customer,
        'form': form,
        'medicines': medicines,
        'purchase_history': purchase_history,
        'total_spent': total_spent,
        'total_visits': total_visits,
    })


@assistant_or_above
def delete_customer(request, id):
    """Delete customer"""
    customer = get_object_or_404(owner_scope_queryset(request, Customer.objects.all(), 'created_by'), id=id)
    name = customer.name
    customer.delete()
    messages.success(request, f'Customer "{name}" deleted successfully!')
    return redirect('customer_list')


@assistant_or_above
def reminder_create(request, customer_id):
    """Create a medicine reminder and send an email to the customer."""
    customer = get_object_or_404(owner_scope_queryset(request, Customer.objects.all(), 'created_by'), id=customer_id)

    if request.method == 'POST':
        form = MedicineReminderForm(request.POST)
        if form.is_valid():
            reminder = form.save(commit=False)
            reminder.customer = customer
            reminder.save()
            reminder.schedule_next()
            reminder.save()

            # ✉️ Send email notification
            email_result = send_reminder_email(customer, reminder)

            # AJAX response
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'id': reminder.id,
                    'email_sent': email_result['success'],
                    'email_error': email_result.get('error')
                })

            if email_result['success']:
                messages.success(request, 'Reminder created and email sent successfully!')
            else:
                messages.warning(request, f"Reminder created, but email could not be sent: {email_result.get('error')}")
            return redirect('edit_customer', customer_id=customer.id)

        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors})

    form = MedicineReminderForm()
    return render(request, 'customer/add_reminder.html', {'form': form, 'customer': customer})


@assistant_or_above
def reminder_detail_json(request, reminder_id):
    r = get_object_or_404(owner_scope_queryset(request, MedicineReminder.objects.select_related('customer__created_by'), 'customer__created_by'), id=reminder_id)
    data = {
        'id': r.id,
        'medicine': r.medicine.id if r.medicine else None,
        'reminder_text': r.reminder_text,
        'period': r.period,
        'custom_days': r.custom_days,
        'send_at': r.send_at.isoformat() if r.send_at else None,
        'next_send': r.next_send.isoformat() if r.next_send else None,
    }
    return JsonResponse(data)


@assistant_or_above
def reminder_update(request, reminder_id):
    r = get_object_or_404(owner_scope_queryset(request, MedicineReminder.objects.select_related('customer__created_by'), 'customer__created_by'), id=reminder_id)
    if request.method == 'POST':
        form = MedicineReminderForm(request.POST, instance=r)
        if form.is_valid():
            r = form.save(commit=False)
            r.schedule_next()
            r.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True})
            messages.success(request, 'Reminder updated.')
            return redirect('edit_customer', customer_id=r.customer.id)
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors})
    # fallback: render edit page
    form = MedicineReminderForm(instance=r)
    return render(request, 'customer/edit_reminder.html', {'form': form, 'reminder': r})


@assistant_or_above
def reminder_delete(request, reminder_id):
    r = get_object_or_404(owner_scope_queryset(request, MedicineReminder.objects.select_related('customer__created_by'), 'customer__created_by'), id=reminder_id)
    customer_id = r.customer.id
    if request.method == 'POST':
        r.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        messages.success(request, 'Reminder deleted.')
        return redirect('edit_customer', customer_id=customer_id)
    return JsonResponse({'success': False, 'error': 'Invalid request'})


@assistant_or_above
def send_reminder_now(request, reminder_id):
    """Trigger manual email sending for a specific customer reminder."""
    reminder = get_object_or_404(owner_scope_queryset(request, MedicineReminder.objects.select_related('customer__created_by'), 'customer__created_by'), id=reminder_id)
    customer = reminder.customer

    email_result = send_reminder_email(customer, reminder)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('format') == 'json':
        if email_result['success']:
            return JsonResponse({'success': True, 'message': f"Reminder email successfully sent to {customer.email}!"})
        else:
            return JsonResponse({'success': False, 'error': email_result.get('error', 'Failed to send email.')})

    if email_result['success']:
        messages.success(request, f"Reminder email successfully sent to {customer.name} ({customer.email})!")
    else:
        messages.error(request, f"Failed to send email to {customer.name}: {email_result.get('error')}")

    return redirect('edit_customer', customer_id=customer.id)


@assistant_or_above
def send_due_reminders_view(request):
    """Bulk send all currently due customer medicine reminders."""
    now = timezone.now()
    due_reminders = owner_scope_queryset(request, MedicineReminder.objects.filter(next_send__lte=now).select_related('customer'), 'customer__created_by')

    sent_count = 0
    errors = []

    for r in due_reminders:
        result = send_reminder_email(r.customer, r)
        if result['success']:
            sent_count += 1
            r.schedule_next(from_dt=now)
            if r.period == 'one_time':
                r.next_send = None
            r.save()
        else:
            errors.append(f"{r.customer.name}: {result.get('error')}")

    msg = f"Processed {sent_count} due reminder(s)."
    if errors:
        msg += f" Encountered {len(errors)} error(s)."

    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('format') == 'json':
        return JsonResponse({
            'success': True,
            'sent_count': sent_count,
            'errors': errors,
            'message': msg
        })

    if errors:
        messages.warning(request, msg)
    else:
        messages.success(request, msg)

    return redirect('customer_list')


# Supplier Views
@pharmacist_or_admin
def supplier_list(request):
    """List all suppliers"""
    search_query = request.GET.get('search', '')
    suppliers = owner_scope_queryset(request, Supplier.objects.prefetch_related('supplied_medicines'), 'created_by')
    
    if search_query:
        suppliers = suppliers.filter(
            Q(name__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    context = {
        'suppliers': suppliers,
        'search_query': search_query,
    }
    return render(request, 'supplier/supplier_list.html', context)


@pharmacist_or_admin
def add_supplier(request):
    """Add new supplier"""
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            supplier = form.save(commit=False)
            supplier.created_by = request.user
            supplier.save()
            messages.success(request, f'Supplier "{supplier.name}" added successfully!')
            return redirect('supplier_list')
    else:
        form = SupplierForm()
    
    return render(request, 'supplier/add_supplier.html', {'form': form})


@pharmacist_or_admin
def edit_supplier(request, id):
    """Edit supplier details"""
    supplier = get_object_or_404(owner_scope_queryset(request, Supplier.objects.all(), 'created_by'), id=id)

    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, f'Supplier "{supplier.name}" updated successfully!')
            return redirect('supplier_list')
    else:
        form = SupplierForm(instance=supplier)

    return render(request, 'supplier/edit_supplier.html', {
        'form': form,
        'supplier': supplier
    })


@pharmacist_or_admin
def delete_supplier(request, id):
    """Delete supplier"""
    supplier = get_object_or_404(owner_scope_queryset(request, Supplier.objects.all(), 'created_by'), id=id)
    name = supplier.name
    supplier.delete()
    messages.success(request, f'Supplier "{name}" deleted successfully!')
    return redirect('supplier_list')


@pharmacist_or_admin
def supplier_profile(request, id):
    """View supplier profile with supplied medicines"""
    supplier = get_object_or_404(owner_scope_queryset(request, Supplier.objects.all(), 'created_by'), id=id)
    
    search_query = request.GET.get('search', '')
    medicines = supplier.supplied_medicines.prefetch_related('batches')
    
    if search_query:
        medicines = medicines.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    context = {
        'supplier': supplier,
        'medicines': medicines,
        'search_query': search_query,
    }
    return render(request, 'supplier/supplier_profile.html', context)


@pharmacist_or_admin
def export_supplier_stock(request, id):
    """Export all stock (batches) supplied by this supplier to an Excel file."""
    import openpyxl
    from django.http import HttpResponse

    supplier = get_object_or_404(owner_scope_queryset(request, Supplier.objects.all(), 'created_by'), id=id)

    # Get all batches for medicines supplied by this supplier
    batches = owner_scope_queryset(request, Batch.objects.filter(medicine__supplier=supplier).select_related('medicine').order_by('-add_date'), 'created_by')

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Stock_{supplier.name[:25]}"

    # Define headers
    headers = ["Medicine Name", "Batch Name", "Added Date", "Expiry Date", "Quantity", "Purchase Price"]
    ws.append(headers)

    # Add data rows
    for b in batches:
        ws.append([
            b.medicine.name if b.medicine else "N/A",
            b.batch_name,
            b.add_date.strftime('%Y-%m-%d') if b.add_date else "N/A",
            b.expiry_date.strftime('%Y-%m-%d') if b.expiry_date else "N/A",
            b.quantity,
            float(b.purchase_price) if b.purchase_price else 0.0
        ])

    # Format the header row
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Create the HTTP response with Excel content type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Supplier_Stock_{supplier.name.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


# API Views
@login_required
def get_medicine_batches(request):
    """API to get batches for a medicine"""
    medicine_id = request.GET.get('medicine_id')
    if not medicine_id:
        return JsonResponse({'batches': []})

    batches = owner_scope_queryset(
        request,
        Batch.objects.filter(
            medicine_id=medicine_id,
            quantity__gt=0
        ).exclude(
            expiry_date__lt=timezone.now().date()
        ).order_by('expiry_date'),
        'created_by'
    )
    
    data = [{
        'id': b.id,
        'batch_name': b.batch_name,
        'quantity': b.quantity,
        'expiry_date': b.expiry_date.strftime('%Y-%m-%d'),
    } for b in batches]
    
    return JsonResponse({'batches': data})


@login_required
def search_medicines(request):
    """API to search medicines"""
    query = request.GET.get('q', '')
    if len(query) < 2:
        return JsonResponse({'medicines': []})

    medicines = owner_scope_queryset(
        request,
        Medicine.objects.filter(
            Q(name__icontains=query) |
            Q(description__icontains=query)
        ),
        'created_by'
    )[:10]
    
    data = [{
        'id': m.id,
        'name': m.name,
        'price': str(m.price),
        'quantity': m.total_quantity,
        'is_active': m.is_active,
    } for m in medicines]
    
    return JsonResponse({'medicines': data})


@login_required
def search_customers(request):
    """API to search customers"""
    query = request.GET.get('q', '')
    if len(query) < 2:
        return JsonResponse({'customers': []})

    customers = owner_scope_queryset(
        request,
        Customer.objects.filter(
            Q(name__icontains=query) |
            Q(email__icontains=query) |
            Q(contact_number__icontains=query)
        ),
        'created_by'
    )[:10]
    
    data = [{
        'id': c.id,
        'name': c.name,
        'email': c.email,
        'contact_number': c.contact_number,
    } for c in customers]
    
    return JsonResponse({'customers': data})


@login_required
def search_suppliers(request):
    """API to search suppliers"""
    query = request.GET.get('q', '')
    if len(query) < 2:
        return JsonResponse({'suppliers': []})

    suppliers = owner_scope_queryset(
        request,
        Supplier.objects.filter(
            Q(name__icontains=query) |
            Q(location__icontains=query)
        ),
        'created_by'
    )[:10]
    
    data = [{
        'id': s.id,
        'name': s.name,
        'location': s.location,
    } for s in suppliers]
    
    return JsonResponse({'suppliers': data})


# Reports Views
@assistant_or_above
def reports_dashboard(request):
    """View for displaying all analytics charts"""
    return render(request, 'reports/reports_dashboard.html')


@assistant_or_above
def profit_loss_report(request):
    """Standalone view for displaying Profit & Loss chart"""
    return render(request, 'reports/profit_loss_report.html')


@assistant_or_above
def payment_mode_report(request):
    """Standalone view for displaying Payment Mode Tracking"""
    return render(request, 'reports/payment_mode_report.html')


@assistant_or_above
def payment_mode_data(request):
    """API for payment mode distribution"""
    filter_type = request.GET.get('filter', 'month')
    today = timezone.now().date()

    start_date = today
    end_date = today

    if filter_type == 'today':
        start_date = today
    elif filter_type == 'week':
        start_date = today - timedelta(days=7)
    elif filter_type == 'month':
        start_date = today - timedelta(days=30)
    elif filter_type == 'year':
        start_date = today - timedelta(days=365)
    elif filter_type == 'custom':
        start_str = request.GET.get('start_date')
        end_str = request.GET.get('end_date')
        if start_str and end_str:
            start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_str, '%Y-%m-%d').date()

    sales = owner_scope_sales(
        request,
        Sale.objects.filter(
            date__date__gte=start_date,
            date__date__lte=end_date,
            status='Completed'
        )
    ).values('payment_method').annotate(
        total_amount=Sum('total_price'),
        count=Count('id')
    ).order_by('-total_amount')
    
    labels = []
    amounts = []
    counts = []
    
    for item in sales:
        labels.append(item['payment_method'])
        amounts.append(float(item['total_amount'] or 0))
        counts.append(item['count'])
        
    return JsonResponse({
        'labels': labels,
        'amounts': amounts,
        'counts': counts
    })



@assistant_or_above
def top_selling_products_data(request):
    """API for top selling products chart"""
    filter_type = request.GET.get('filter', 'month')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    today = timezone.now().date()

    if filter_type == 'today':
        start = today
        end = today
    elif filter_type == 'week':
        start = today - timedelta(days=7)
        end = today
    elif filter_type == 'month':
        start = today - timedelta(days=30)
        end = today
    elif filter_type == 'custom' and start_date and end_date:
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        start = today - timedelta(days=30)
        end = today

    scoped_sales = owner_scope_sales(request, Sale.objects.filter(date__date__gte=start, date__date__lte=end))
    top_products = (
        SaleItem.objects
        .filter(sale__in=scoped_sales)
        .values('medicine__name')
        .annotate(total_sold=Sum('quantity'))
        .order_by('-total_sold')[:10]
    )
    
    data = {
        'labels': [p['medicine__name'] for p in top_products],
        'data': [p['total_sold'] for p in top_products],
    }
    
    return JsonResponse(data)


@assistant_or_above
def total_sales_over_time_data(request):
    """API for sales over time chart"""
    filter_type = request.GET.get('filter', 'month')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    today = timezone.now().date()

    if filter_type == 'today':
        start = today
        end = today
    elif filter_type == 'week':
        start = today - timedelta(days=7)
        end = today
    elif filter_type == 'month':
        start = today - timedelta(days=30)
        end = today
    elif filter_type == 'custom' and start_date and end_date:
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        start = today - timedelta(days=30)
        end = today

    scoped_sales = owner_scope_sales(request, Sale.objects.filter(date__date__gte=start, date__date__lte=end))
    sales_data = (
        scoped_sales
        .annotate(sale_date=TruncDate('date'))
        .values('sale_date')
        .annotate(total=Sum('total_price'))
        .order_by('sale_date')
    )
    
    data = {
        'labels': [s['sale_date'].strftime('%Y-%m-%d') for s in sales_data],
        'data': [float(s['total']) for s in sales_data],
    }
    
    return JsonResponse(data)


@assistant_or_above
def profit_loss_data(request):
    """API for profit/loss chart"""
    from django.db.models import F
    filter_type = request.GET.get('filter', 'month')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    today = timezone.now().date()

    if filter_type == 'today':
        start = today
        end = today
    elif filter_type == 'week':
        start = today - timedelta(days=7)
        end = today
    elif filter_type == 'month':
        start = today - timedelta(days=30)
        end = today
    elif filter_type == 'custom' and start_date and end_date:
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        start = today - timedelta(days=30)
        end = today

    scoped_sales = owner_scope_sales(request, Sale.objects.filter(date__date__gte=start, date__date__lte=end))
    completed_scoped_sales = scoped_sales.filter(status='Completed')
    sales_items = (
        SaleItem.objects
        .filter(sale__in=completed_scoped_sales)
        .annotate(sale_date=TruncDate('sale__date'))
        .values('sale_date')
        .annotate(
            revenue=Sum(F('price') * F('quantity')),
            cost=Sum(F('cost_price') * F('quantity'))
        )
        .order_by('sale_date')
    )

    sales_discounts = (
        completed_scoped_sales
        .annotate(sale_date=TruncDate('date'))
        .values('sale_date')
        .annotate(total_discount=Sum('discount'))
        .order_by('sale_date')
    )
    
    discount_dict = {s['sale_date'].strftime('%Y-%m-%d'): float(s['total_discount'] or 0) for s in sales_discounts}
    
    labels = []
    revenue_data = []
    profit_data = []
    
    for item in sales_items:
        date_str = item['sale_date'].strftime('%Y-%m-%d')
        labels.append(date_str)
        
        rev = float(item['revenue'] or 0)
        cost = float(item['cost'] or 0)
        disc = discount_dict.get(date_str, 0)
        
        net_revenue = rev - disc
        profit = net_revenue - cost
        
        revenue_data.append(net_revenue)
        profit_data.append(profit)
        
    data = {
        'labels': labels,
        'revenue': revenue_data,
        'profit': profit_data,
    }
    
    return JsonResponse(data)


@assistant_or_above
def customer_registrations_data(request):
    """API for customer registrations chart"""
    filter_type = request.GET.get('filter', 'month')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    today = timezone.now().date()

    if filter_type == 'today':
        start = today
        end = today
    elif filter_type == 'week':
        start = today - timedelta(days=7)
        end = today
    elif filter_type == 'month':
        start = today - timedelta(days=30)
        end = today
    elif filter_type == 'custom' and start_date and end_date:
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        start = today - timedelta(days=30)
        end = today

    registrations = (
        owner_scope_queryset(request, Customer.objects.all(), 'created_by')
        .filter(created_at__date__gte=start, created_at__date__lte=end)
        .annotate(reg_date=TruncDate('created_at'))
        .values('reg_date')
        .annotate(count=Count('id'))
        .order_by('reg_date')
    )
    
    data = {
        'labels': [r['reg_date'].strftime('%Y-%m-%d') for r in registrations],
        'data': [r['count'] for r in registrations],
    }
    
    return JsonResponse(data)


@assistant_or_above
def inventory_additions_data(request):
    """API for inventory additions chart"""
    filter_type = request.GET.get('filter', 'month')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    today = timezone.now().date()

    if filter_type == 'today':
        start = today
        end = today
    elif filter_type == 'week':
        start = today - timedelta(days=7)
        end = today
    elif filter_type == 'month':
        start = today - timedelta(days=30)
        end = today
    elif filter_type == 'custom' and start_date and end_date:
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        start = today - timedelta(days=30)
        end = today

    additions = (
        owner_scope_queryset(request, Medicine.objects.all(), 'created_by')
        .filter(created_at__date__gte=start, created_at__date__lte=end)
        .annotate(add_date=TruncDate('created_at'))
        .values('add_date')
        .annotate(count=Count('id'))
        .order_by('add_date')
    )
    
    data = {
        'labels': [a['add_date'].strftime('%Y-%m-%d') for a in additions],
        'data': [a['count'] for a in additions],
    }
    
    return JsonResponse(data)

# Stock Notification View
@pharmacist_or_admin
@require_POST
def notify_stock_available(request, medicine_id):
    """
    Send a stock-availability email notification to all permanent customers.
    Accessible via AJAX (returns JSON) or regular POST (redirects with flash message).
    Restricted to pharmacists and admins.
    """
    from .utils.email_utils import send_stock_available_email

    medicine = get_object_or_404(Medicine, id=medicine_id)

    # Only notify if medicine actually has stock
    if medicine.total_quantity <= 0:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': f'"{medicine.name}" has no stock. Cannot send notification.',
            }, status=400)
        messages.warning(request, f'"{medicine.name}" has no stock. Notification not sent.')
        return redirect('medicine_list')

    # Fetch all permanent customers who have a valid email within the same user-scoped dataset
    permanent_customers = owner_scope_queryset(request, Customer.objects.filter(is_permanent=True).exclude(email=''), 'created_by')

    if not permanent_customers.exists():
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': 'No permanent customers found. Please mark customers as permanent first.',
            }, status=400)
        messages.warning(request, 'No permanent customers found. Please mark customers as permanent first.')
        return redirect('medicine_list')

    result = send_stock_available_email(medicine, permanent_customers)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'sent_count': result['sent'],
            'errors': result['errors'],
            'message': (
                f'✅ Notification sent to {result["sent"]} permanent customer(s)!'
                + (f' ({len(result["errors"])} failed)' if result['errors'] else '')
            ),
        })

    if result['errors']:
        messages.warning(
            request,
            f'Sent {result["sent"]} notification(s). Failed: {"; ".join(result["errors"])}'
        )
    else:
        messages.success(
            request,
            f'✅ Stock notification for "{medicine.name}" sent to {result["sent"]} permanent customer(s)!'
        )
    return redirect('medicine_list')


# Excel Export Views
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


@assistant_or_above
def export_top_products_excel(request):
    """Export top products to Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse('Excel export not available. Install openpyxl.', status=500)

    filter_type = request.GET.get('filter', 'month')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    today = timezone.now().date()
    if filter_type == 'today':
        start = today
        end = today
    elif filter_type == 'week':
        start = today - timedelta(days=7)
        end = today
    elif filter_type == 'month':
        start = today - timedelta(days=30)
        end = today
    elif filter_type == 'custom' and start_date and end_date:
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        start = today - timedelta(days=30)
        end = today

    scoped_sales = owner_scope_sales(request, Sale.objects.filter(date__date__gte=start, date__date__lte=end))
    # Get data
    top_products = (
        SaleItem.objects
        .filter(sale__in=scoped_sales)
        .values('medicine__name')
        .annotate(
            total_sold=Sum('quantity'),
            total_revenue=Sum('price')
        )
        .order_by('-total_sold')[:10]
    )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Top Selling Products"
    
    # Header style
    header_fill = PatternFill(start_color="0891B2", end_color="0891B2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ['Rank', 'Product Name', 'Units Sold', 'Revenue (₹)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for idx, product in enumerate(top_products, 1):
        ws.cell(row=idx+1, column=1, value=idx)
        ws.cell(row=idx+1, column=2, value=product['medicine__name'])
        ws.cell(row=idx+1, column=3, value=product['total_sold'])
        ws.cell(row=idx+1, column=4, value=float(product['total_revenue']))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="top_products_{start}_{end}.xlsx"'
    wb.save(response)
    
    return response


@assistant_or_above
def export_sales_over_time_excel(request):
    """Export sales over time to Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse('Excel export not available. Install openpyxl.', status=500)

    filter_type = request.GET.get('filter', 'month')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    today = timezone.now().date()
    if filter_type == 'today':
        start = today
        end = today
    elif filter_type == 'week':
        start = today - timedelta(days=7)
        end = today
    elif filter_type == 'month':
        start = today - timedelta(days=30)
        end = today
    elif filter_type == 'custom' and start_date and end_date:
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        start = today - timedelta(days=30)
        end = today

    scoped_sales = owner_scope_sales(request, Sale.objects.filter(date__date__gte=start, date__date__lte=end))
    # Get data
    sales_by_date = (
        scoped_sales
        .annotate(sale_date=TruncDate('date'))
        .values('sale_date')
        .annotate(
            total_sales=Sum('total_price'),
            total_discount=Sum('discount'),
            count=Count('id')
        )
        .order_by('sale_date')
    )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Over Time"
    
    # Header style
    header_fill = PatternFill(start_color="0891B2", end_color="0891B2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ['Date', 'Number of Sales', 'Total Revenue (₹)', 'Total Discount (₹)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for idx, item in enumerate(sales_by_date, 1):
        ws.cell(row=idx+1, column=1, value=item['sale_date'].strftime('%Y-%m-%d'))
        ws.cell(row=idx+1, column=2, value=item['count'])
        ws.cell(row=idx+1, column=3, value=float(item['total_sales']))
        ws.cell(row=idx+1, column=4, value=float(item['total_discount'] or 0))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="sales_over_time_{start}_{end}.xlsx"'
    wb.save(response)
    
    return response


@assistant_or_above
def export_customer_registrations_excel(request):
    """Export customer registrations to Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse('Excel export not available. Install openpyxl.', status=500)

    filter_type = request.GET.get('filter', 'month')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    today = timezone.now().date()
    if filter_type == 'today':
        start = today
        end = today
    elif filter_type == 'week':
        start = today - timedelta(days=7)
        end = today
    elif filter_type == 'month':
        start = today - timedelta(days=30)
        end = today
    elif filter_type == 'custom' and start_date and end_date:
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        start = today - timedelta(days=30)
        end = today

    scoped_customers = owner_scope_queryset(request, Customer.objects.all(), 'created_by')
    # Get data
    registrations = (
        scoped_customers
        .filter(created_at__date__gte=start, created_at__date__lte=end)
        .annotate(reg_date=TruncDate('created_at'))
        .values('reg_date')
        .annotate(count=Count('id'))
        .order_by('reg_date')
    )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Registrations"
    
    # Header style
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ['Date', 'New Customers', 'Cumulative Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data with cumulative count
    cumulative = 0
    for idx, item in enumerate(registrations, 1):
        cumulative += item['count']
        ws.cell(row=idx+1, column=1, value=item['reg_date'].strftime('%Y-%m-%d'))
        ws.cell(row=idx+1, column=2, value=item['count'])
        ws.cell(row=idx+1, column=3, value=cumulative)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="customer_registrations_{start}_{end}.xlsx"'
    wb.save(response)
    
    return response


@assistant_or_above
def export_inventory_additions_excel(request):
    """Export inventory additions to Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse('Excel export not available. Install openpyxl.', status=500)

    filter_type = request.GET.get('filter', 'month')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    today = timezone.now().date()
    if filter_type == 'today':
        start = today
        end = today
    elif filter_type == 'week':
        start = today - timedelta(days=7)
        end = today
    elif filter_type == 'month':
        start = today - timedelta(days=30)
        end = today
    elif filter_type == 'custom' and start_date and end_date:
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        start = today - timedelta(days=30)
        end = today

    scoped_medicines = owner_scope_queryset(request, Medicine.objects.all(), 'created_by')
    # Get data
    additions = (
        scoped_medicines
        .filter(created_at__date__gte=start, created_at__date__lte=end)
        .annotate(add_date=TruncDate('created_at'))
        .values('add_date')
        .annotate(count=Count('id'))
        .order_by('add_date')
    )

    # Also get detailed list of medicines added
    medicines_added = (
        scoped_medicines
        .filter(created_at__date__gte=start, created_at__date__lte=end)
        .values('name', 'created_at', 'supplier__name', 'price')
        .order_by('created_at')
    )
    
    # Create workbook
    wb = Workbook()
    
    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ['Date', 'Products Added']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for idx, item in enumerate(additions, 1):
        ws1.cell(row=idx+1, column=1, value=item['add_date'].strftime('%Y-%m-%d'))
        ws1.cell(row=idx+1, column=2, value=item['count'])
    
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 18
    
    # Sheet 2: Detailed List
    ws2 = wb.create_sheet(title="Detailed List")
    
    headers2 = ['Date Added', 'Medicine Name', 'Supplier', 'Price (₹)']
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for idx, item in enumerate(medicines_added, 1):
        ws2.cell(row=idx+1, column=1, value=item['created_at'].strftime('%Y-%m-%d %H:%M'))
        ws2.cell(row=idx+1, column=2, value=item['name'])
        ws2.cell(row=idx+1, column=3, value=item['supplier__name'] or 'N/A')
        ws2.cell(row=idx+1, column=4, value=float(item['price']))
    
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 15
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="inventory_additions_{start}_{end}.xlsx"'
    wb.save(response)
    
    return response
@pharmacist_or_admin
def category_list(request):
    """List all medicine categories"""
    query = request.GET.get('q', '')
    categories = owner_scope_queryset(request, Category.objects.all(), 'created_by')
    if query:
        categories = categories.filter(
            Q(name__icontains=query) | Q(description__icontains=query)
        ).annotate(medicine_count=Count('medicines')).order_by('name')
    else:
        categories = categories.annotate(medicine_count=Count('medicines')).order_by('name')

    return render(request, 'category/category_list.html', {
        'categories': categories,
        'search_query': query
    })

@pharmacist_or_admin
def add_category(request):
    """Add a new category"""
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save(commit=False)
            category.created_by = request.user
            category.save()
            messages.success(request, f'Category {category.name} added successfully!')
            return redirect('category_list')
    else:
        form = CategoryForm()

    return render(request, 'category/add_category.html', {'form': form})

@pharmacist_or_admin
def edit_category(request, id):
    """Edit an existing category"""
    category = get_object_or_404(Category, id=id)
    
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Category {category.name} updated successfully!')
            return redirect('category_list')
    else:
        form = CategoryForm(instance=category)
        
    return render(request, 'category/edit_category.html', {
        'form': form,
        'category': category
    })

@admin_required
def delete_category(request, id):
    """Delete a category"""
    category = get_object_or_404(Category, id=id)
    
    if request.method == 'POST':
        if category.medicines.exists():
            messages.error(request, 'Cannot delete category because it has associated medicines.')
        else:
            name = category.name
            category.delete()
            messages.success(request, f'Category {name} deleted successfully!')
        return redirect('category_list')
        
    return render(request, 'category/delete_category.html', {'category': category})


# =============================================================================
# MediBot - AI Pharmacy Assistant Engine (Rebuilt from Scratch)
# =============================================================================
import difflib
import html
import json
import logging
import re
import urllib.error
import urllib.request
from datetime import timedelta
from decimal import Decimal
from django.conf import settings as django_settings
from django.http import JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_POST

logger = logging.getLogger(__name__)

# Medical Guidance & Indications Dictionary for Pharmacy Queries
MEDIBOT_MEDICAL_KNOWLEDGE = {
    'paracetamol': {
        'generic_name': 'Paracetamol (Acetaminophen / APAP)',
        'drug_class': 'Analgesic & Antipyretic (Non-opioid pain reliever & fever reducer)',
        'uses': 'Relief of mild-to-moderate pain such as headaches, migraines, dental pain, toothaches, body aches, muscle aches, post-immunization pyrexia, backaches, osteoarthritis pain, and fever reduction.',
        'side_effects': 'Generally well-tolerated at therapeutic doses.\n• **Common/Mild:** Mild nausea, dyspepsia, skin rash or allergic reactions.\n• **Severe / Overdose Risks (Disadvantages):** Severe acute liver failure / hepatotoxicity (liver necrosis) is the primary disadvantage if exceeded beyond 4000mg/day or combined with alcohol. Long-term high doses can also increase risk of renal impairment or thrombocytopenia.',
        'dosage': '• **Adults & Adolescents (>50kg):** 500mg to 1000mg orally every 4 to 6 hours as needed.\n• **Maximum Daily Limit:** 4000mg (4 grams) in 24 hours.\n• **Dosing Interval:** Never take doses closer than 4 hours apart.\n• **Children:** 10–15 mg/kg per dose every 4–6 hours (max 60 mg/kg/day under pediatric supervision).',
        'pregnancy': 'Considered the safest first-line analgesic and antipyretic during all trimesters of pregnancy and while breastfeeding when taken at the lowest effective dose for the shortest duration necessary. Always consult an obstetrician.',
        'precautions': '• Avoid consuming alcohol as it significantly elevates liver toxicity risks.\n• Check labels on cough/cold combination products to prevent accidental paracetamol overdose.\n• Use with caution in patients with chronic liver disease, severe renal impairment, or chronic malnutrition.\n• **Drug Interactions:** May interact with Warfarin (prolonged regular use increases INR/bleeding risk), Isoniazid, and Carbamazepine.',
    },
    'azithromycin': {
        'generic_name': 'Azithromycin (Azithral / Zithromax)',
        'drug_class': 'Macrolide Antibiotic',
        'uses': 'Treatment of mild-to-moderate bacterial infections including upper and lower respiratory tract infections (community-acquired pneumonia, acute bronchitis, tonsillitis, pharyngitis, sinusitis), skin/soft-tissue infections, ear infections (otitis media), and certain sexually transmitted infections (Chlamydia, Gonorrhea).',
        'side_effects': '• **Common (Disadvantages):** Gastrointestinal upset (diarrhea, nausea, abdominal pain/cramping, vomiting), flatulence, headache, dizziness, altered taste.\n• **Severe / Rare Risks:** QT interval prolongation (cardiac arrhythmia), cholestatic jaundice / hepatic dysfunction, severe Clostridium difficile-associated diarrhea (pseudomembranous colitis), allergic angioedema or Stevens-Johnson syndrome.',
        'dosage': '• **Standard Adult Regimen:** 500mg once daily as a single dose on Day 1, followed by 250mg once daily on Days 2–5 (total 1.5g course); OR 500mg once daily for 3 consecutive days depending on indication.\n• Take with or without food (taking with food reduces GI discomfort). Swallow tablets whole with a glass of water.',
        'pregnancy': 'Category B. Generally considered acceptable during pregnancy and lactation when clearly indicated and prescribed by a doctor, with no established evidence of teratogenicity in clinical studies.',
        'precautions': '• Complete the full prescribed course even if symptoms improve early to prevent bacterial resistance.\n• Avoid co-administration with aluminium- and magnesium-containing antacids (space by at least 2 hours).\n• Caution in patients with known cardiac rhythm disorders (QT prolongation, bradycardia) or severe liver impairment.\n• **Interactions:** Anticoagulants (Warfarin), Digoxin, Colchicine, and Cyclosporine.',
    },
    'amoxicillin': {
        'generic_name': 'Amoxicillin (Mox / Amoxil)',
        'drug_class': 'Broad-Spectrum Aminopenicillin Antibiotic',
        'uses': 'Treatment of susceptible bacterial infections including acute otitis media (ear infection), streptococcal pharyngitis (strep throat), pneumonia, bronchitis, sinusitis, urinary tract infections (UTIs), skin infections, and as part of combination therapy for Helicobacter pylori eradication in peptic ulcers.',
        'side_effects': '• **Common (Disadvantages):** Diarrhea, nausea, vomiting, mild skin rash, oral or vaginal candidiasis (thrush).\n• **Severe Risks:** Severe hypersensitivity / anaphylaxis in penicillin-allergic patients, severe pseudomembranous colitis (C. difficile diarrhea), interstitial nephritis, elevated liver enzymes.',
        'dosage': '• **Standard Adult Dosage:** 250mg to 500mg every 8 hours (three times daily) OR 500mg to 875mg every 12 hours (twice daily) for 7 to 10 days depending on infection severity.\n• **Administration:** Can be taken with or without meals; taking with food helps prevent stomach upset.',
        'pregnancy': 'Category B. Widely regarded as safe and commonly prescribed during pregnancy and breastfeeding for susceptible bacterial infections under physician supervision.',
        'precautions': '• **Strict Contraindication:** Known allergy/hypersensitivity to penicillins or beta-lactam antibiotics.\n• Always complete the full course to prevent recurrence and antibiotic resistance.\n• **Interactions:** Methotrexate (increased toxicity), Allopurinol (increased rash incidence), Oral Contraceptives (may reduce efficacy), and Probenecid.',
    },
    'cetirizine': {
        'generic_name': 'Cetirizine (Cetzine / Okacet / Zyrtec)',
        'drug_class': 'Second-Generation Antihistamine (H1 Receptor Antagonist)',
        'uses': 'Relief of allergy symptoms associated with seasonal and perennial allergic rhinitis (sneezing, rhinorrhea, nasal congestion, itchy/watery eyes) and treatment of chronic idiopathic urticaria (itchy hives and skin rashes).',
        'side_effects': '• **Common (Disadvantages):** Mild drowsiness/sedation, dry mouth, fatigue, dizziness, headache, gastrointestinal discomfort.\n• **Rare:** Tachycardia, urinary retention, allergic bronchospasm.',
        'dosage': '• **Adults & Children >12 years:** 5mg to 10mg once daily, preferably in the evening/bedtime.\n• **Elderly & Renal Impairment:** 5mg once daily recommended.',
        'pregnancy': 'Category B. Generally safe when prescribed by a doctor, but non-drug measures or first-line alternatives are preferred where appropriate.',
        'precautions': '• Avoid driving, operating heavy machinery, or drinking alcohol while taking cetirizine due to potential additive sedative effects.\n• Caution in patients with prostatic hypertrophy or severe renal impairment.',
    },
    'amlodipine': {
        'generic_name': 'Amlodipine (Amlong / Norvasc)',
        'drug_class': 'Dihydropyridine Calcium Channel Blocker (CCB)',
        'uses': 'Management of essential hypertension (high blood pressure) and treatment of chronic stable angina pectoris and vasospastic (Prinzmetal\'s) angina.',
        'side_effects': '• **Common (Disadvantages):** Peripheral edema (swelling of ankles/feet), flushing, dizziness, palpitations, fatigue, headache, nausea.\n• **Severe/Rare:** Excessive hypotension (low BP), gingival hyperplasia, bradycardia/tachycardia.',
        'dosage': '• **Initial Adult Dose:** 5mg once daily orally, which may be increased to a maximum of 10mg once daily after 1–2 weeks based on clinical response.\n• Can be taken morning or evening, with or without food.',
        'pregnancy': 'Category C. Use only if potential benefit justifies potential risk to the fetus; consult a cardiologist / obstetrician.',
        'precautions': '• Do not abruptly discontinue taking amlodipine without medical supervision.\n• Caution in patients with severe aortic stenosis or advanced heart failure.\n• Avoid excessive grapefruit juice as it may increase amlodipine blood concentrations.',
    },
    'metformin': {
        'generic_name': 'Metformin Hydrochloride (Glycomet / Glucophage)',
        'drug_class': 'Biguanide Oral Hypoglycemic / Antidiabetic Agent',
        'uses': 'First-line pharmacological management of Type 2 Diabetes Mellitus to lower blood glucose and improve insulin sensitivity. Also used off-label in Polycystic Ovary Syndrome (PCOS).',
        'side_effects': '• **Common (Disadvantages):** Gastrointestinal disturbances (metallic taste, nausea, abdominal bloating, diarrhea, flatulence, anorexia), long-term Vitamin B12 deficiency.\n• **Severe/Rare:** Lactic acidosis (rare but life-threatening emergency, particularly in renal/hepatic impairment).',
        'dosage': '• **Initial Adult Dose:** 500mg orally once or twice daily with meals, gradually titrated to 850mg–1000mg twice daily (maximum 2000mg–2550mg/day).\n• **Administration:** Always take with or immediately after meals to minimize gastrointestinal adverse effects.',
        'pregnancy': 'Category B. Frequently utilized in gestational diabetes under specialist guidance.',
        'precautions': '• **Contraindications:** Severe renal impairment (eGFR < 30 mL/min), acute metabolic acidosis, severe hypoxia, congestive heart failure.\n• Temporarily discontinue prior to iodinated radiocontrast procedures.',
    },
    'pantoprazole': {
        'generic_name': 'Pantoprazole (Pantocid / Pantodac / Pan 40 / Protonix)',
        'drug_class': 'Proton Pump Inhibitor (PPI)',
        'uses': 'Treatment and symptomatic relief of Gastroesophageal Reflux Disease (GERD), erosive esophagitis, gastric and duodenal ulcers, Zollinger-Ellison syndrome, and prevention of NSAID-induced ulcers.',
        'side_effects': '• **Common (Disadvantages):** Headache, diarrhea, nausea, abdominal pain, constipation, flatulence, dizziness.\n• **Long-Term Risks:** Hypomagnesemia, Vitamin B12 malabsorption, increased risk of bone fractures (osteoporosis), Clostridium difficile infection, interstitial nephritis.',
        'dosage': '• **Standard Dose:** 40mg once daily in the morning, 30 to 60 minutes before breakfast.\n• Swallow tablet whole; do not crush, chew, or split delayed-release tablets.',
        'pregnancy': 'Category B. Generally acceptable during pregnancy when clinically warranted under medical supervision.',
        'precautions': '• Long-term continuous use (>1 year) requires periodic monitoring of bone density, magnesium, and B12 levels.\n• May reduce absorption of drugs requiring acidic gastric pH (e.g. Ketoconazole, Iron supplements).',
    },
    'ibuprofen': {
        'generic_name': 'Ibuprofen (Brufen / Advil / Motrin)',
        'drug_class': 'Non-Steroidal Anti-Inflammatory Drug (NSAID)',
        'uses': 'Relief of mild-to-moderate pain, acute inflammation, joint swelling in rheumatoid arthritis, osteoarthritis, dysmenorrhea (menstrual cramps), dental pain, and fever reduction.',
        'side_effects': '• **Common (Disadvantages):** Dyspepsia, heartburn, epigastric pain, nausea, gastrointestinal ulceration/bleeding, fluid retention, headache.\n• **Severe Risks:** GI bleeding/perforation, cardiovascular events (myocardial infarction/stroke with chronic high doses), acute renal dysfunction, bronchospasm in aspirin-sensitive asthmatics.',
        'dosage': '• **Adults:** 200mg to 400mg every 4 to 6 hours as needed with or after food. Maximum OTC limit: 1200mg/day; Prescription max: 2400mg/day under supervision.\n• Always take with food or milk to minimize gastric mucosal irritation.',
        'pregnancy': 'Category C (1st/2nd trimester); Category D / Contraindicated in 3rd trimester (causes premature closure of fetal ductus arteriosus and oligohydramnios). Avoid in pregnancy.',
        'precautions': '• Avoid in patients with active peptic ulcer disease, severe heart failure, or history of NSAID-induced asthma.\n• Caution with antihypertensive medications and anticoagulants.',
    },
    'omeprazole': {
        'generic_name': 'Omeprazole (Omez / Prilosec)',
        'drug_class': 'Proton Pump Inhibitor (PPI)',
        'uses': 'Treatment of heartburn, acid reflux, GERD, gastric and duodenal ulcers, and H. pylori eradication therapy.',
        'side_effects': 'Headache, diarrhea, stomach pain, flatulence, nausea, Vitamin B12 deficiency on long-term use.',
        'dosage': '20mg to 40mg once daily in the morning before meals for 4 to 8 weeks.',
        'pregnancy': 'Category C. Consult a doctor before use during pregnancy.',
        'precautions': 'Do not chew or crush capsules. Take 30–60 mins prior to a meal.',
    },
    'telmisartan': {
        'generic_name': 'Telmisartan (Telma / Micardis)',
        'drug_class': 'Angiotensin II Receptor Blocker (ARB)',
        'uses': 'Treatment of essential hypertension (high blood pressure) and reduction of cardiovascular risk in vulnerable adults.',
        'side_effects': 'Dizziness, upper respiratory tract infection, back pain, sinus congestion, hyperkalemia (high potassium).',
        'dosage': '40mg once daily, can be increased to 80mg once daily if needed.',
        'pregnancy': 'Strictly CONTRAINDICATED in 2nd and 3rd trimesters of pregnancy due to risk of fetal harm.',
        'precautions': 'Monitor serum potassium and renal function periodically. Avoid potassium supplements unless advised.',
    },
    'atorvastatin': {
        'generic_name': 'Atorvastatin (Atorva / Lipitor)',
        'drug_class': 'HMG-CoA Reductase Inhibitor (Statin)',
        'uses': 'Lowering LDL cholesterol, total cholesterol, and triglycerides, while raising HDL; prevention of cardiovascular disease.',
        'side_effects': 'Myalgia (muscle pain), headache, mild GI upset, elevated liver enzymes, rare risk of rhabdomyolysis.',
        'dosage': '10mg to 80mg once daily taken in the evening or bedtime.',
        'pregnancy': 'CONTRAINDICATED in pregnancy and lactation (Category X).',
        'precautions': 'Report unexplained muscle pain or weakness immediately. Avoid grapefruit juice.',
    },
    'ciprofloxacin': {
        'generic_name': 'Ciprofloxacin (Cifran / Cipro)',
        'drug_class': 'Fluoroquinolone Antibiotic',
        'uses': 'Treatment of bacterial infections including complicated urinary tract infections (UTIs), infectious diarrhea, bone/joint infections, and respiratory infections.',
        'side_effects': 'Nausea, diarrhea, tendonitis / tendon rupture risk (especially Achilles tendon), QT prolongation, CNS stimulation.',
        'dosage': '250mg to 750mg every 12 hours depending on infection severity.',
        'pregnancy': 'Category C. Avoid unless no safer alternative is available.',
        'precautions': 'Drink plenty of water. Avoid taking with dairy, calcium, or antacids (space by 2 hours).',
    }
}

# Common Brand Names to Generic Drug Mapping
MEDIBOT_BRAND_ALIASES = {
    'dolo': 'Paracetamol',
    'crocin': 'Paracetamol',
    'calpol': 'Paracetamol',
    'mox': 'Amoxicillin',
    'azithral': 'Azithromycin',
    'amlong': 'Amlodipine',
    'omez': 'Omeprazole',
    'pantocid': 'Pantoprazole',
    'pantodac': 'Pantoprazole',
    'glycomet': 'Metformin',
    'voveran': 'Diclofenac',
    'brufen': 'Ibuprofen',
    'asthalin': 'Salbutamol',
    'cetzine': 'Cetirizine',
    'okacet': 'Cetirizine',
    'budecort': 'Budesonide',
    'telma': 'Telmisartan',
    'metrogyl': 'Metronidazole',
    'pan 40': 'Pantoprazole',
    'pan-40': 'Pantoprazole',
    'pan d': 'Pantoprazole',
}


def _clean_user_prompt(text):
    """Normalize and clean user query string for intent matching."""
    if not text:
        return ""
    q = text.strip().lower()

    # Split common concatenated phrases
    replacements = [
        (r'\bwhatis\b', 'what is'),
        (r'\bwhatare\b', 'what are'),
        (r'\bhowmuch\b', 'how much'),
        (r'\bhowmany\b', 'how many'),
        (r'\bstockof\b', 'stock of'),
        (r'\bpriceof\b', 'price of'),
        (r'\bcostof\b', 'cost of'),
        (r'\bexpiryof\b', 'expiry of'),
        (r'\bexpireof\b', 'expiry of'),
        (r'\bexpiringin\b', 'expiring in'),
        (r'\bcheckstock\b', 'check stock'),
        (r'\bdoyouhave\b', 'do you have'),
        (r'\bavailablestock\b', 'available stock'),
        (r'\btotalstock\b', 'total stock'),
        (r'\blowstock\b', 'low stock'),
    ]
    for pat, rep in replacements:
        q = re.sub(pat, rep, q, flags=re.IGNORECASE)

    q = re.sub(r'[^\w\s\-\.%]', ' ', q)
    return re.sub(r'\s+', ' ', q).strip()


def _classify_query_intent(prompt, session=None):
    """
    Classify user query into:
    - MEDICAL_INFO (sub_aspect: SIDE_EFFECTS, PREGNANCY, DOSAGE, PRECAUTIONS, USES, GENERAL_MEDICAL)
    - CUSTOMER_INFO
    - DEMAND_FORECAST
    - PROFIT_MARGIN
    - PATIENT_REFILL
    - STOCK
    - PRICE
    - EXPIRY
    - SALES_REVENUE
    - SUPPLIERS
    - CATEGORIES
    - HELP
    - STOCK_PRICE_DETAIL
    - GENERAL_QUERY
    """
    raw = (prompt or '').strip().lower()
    clean = _clean_user_prompt(raw)

    # Contextual memory resolution (e.g. "iska price", "iski expiry", "it", "this medicine")
    session_med_key = session.get('medibot_last_med_key') if session else None
    has_pronoun_ref = bool(re.search(r'\b(iska|iski|isko|it|its|this medicine|this drug|this item|inka|inki)\b', clean, flags=re.IGNORECASE))

    # 1. Match medicine keyword if mentioned
    matched_med_key = None
    for k in MEDIBOT_MEDICAL_KNOWLEDGE.keys():
        if re.search(r'\b' + re.escape(k) + r'\b', clean, flags=re.IGNORECASE):
            matched_med_key = k
            break
    if not matched_med_key:
        for brand_k, gen_v in MEDIBOT_BRAND_ALIASES.items():
            if re.search(r'\b' + re.escape(brand_k) + r'\b', clean, flags=re.IGNORECASE):
                matched_med_key = gen_v.lower()
                break

    if not matched_med_key and has_pronoun_ref and session_med_key:
        matched_med_key = session_med_key

    # 2. Check for specific medical sub-aspects FIRST
    is_side_effects = bool(re.search(r'\b(disadvantage|disadvantages|side effect|side effects|adverse effect|adverse effects|adverse reaction|adverse reactions|harm|harmful|danger|dangers|risk|risks|toxicity|toxic|overdose|negative effect|negative effects|drawback|drawbacks|complications|problem|problems)\b', clean, flags=re.IGNORECASE))
    
    is_pregnancy = bool(re.search(r'\b(pregnancy|pregnant|safe in pregnancy|safe during pregnancy|safe for pregnant|breastfeeding|lactation|nursing|baby|infant|fetus|trimester)\b', clean, flags=re.IGNORECASE))
    
    is_dosage = bool(re.search(r'\b(dosage|dose|doses|how much to take|how to take|how many tablets|how many pills|how many times|daily limit|maximum dose|max dose|when to take|before food|after food|frequency|administration)\b', clean, flags=re.IGNORECASE))
    
    is_precautions = bool(re.search(r'\b(precaution|precautions|warning|warnings|contraindication|contraindications|interaction|interactions|mix with|with alcohol|who should not take|avoid with|safety warning)\b', clean, flags=re.IGNORECASE))
    
    is_uses = bool(re.search(r'\b(use of|uses of|use|uses|used for|what is it for|what is it used for|why use|why take|benefit|benefits|indication|indications|prescribed for|treat|treatment|cure|action of|purpose of)\b', clean, flags=re.IGNORECASE))
    
    is_general_medical = bool(re.search(r'\b(what is|what are|tell me about|information about|info about|about|how does|how it works)\s+[a-zA-Z0-9]', clean, flags=re.IGNORECASE)) and not bool(re.search(r'\b(stock|price|cost|expiry|sales|profit|margin)\b', clean, flags=re.IGNORECASE))

    if is_side_effects:
        return {'intent': 'MEDICAL_INFO', 'sub_aspect': 'SIDE_EFFECTS', 'med_key': matched_med_key}
    if is_pregnancy:
        return {'intent': 'MEDICAL_INFO', 'sub_aspect': 'PREGNANCY', 'med_key': matched_med_key}
    if is_dosage:
        return {'intent': 'MEDICAL_INFO', 'sub_aspect': 'DOSAGE', 'med_key': matched_med_key}
    if is_precautions:
        return {'intent': 'MEDICAL_INFO', 'sub_aspect': 'PRECAUTIONS', 'med_key': matched_med_key}
    if is_uses:
        return {'intent': 'MEDICAL_INFO', 'sub_aspect': 'USES', 'med_key': matched_med_key}
    if is_general_medical and matched_med_key:
        return {'intent': 'MEDICAL_INFO', 'sub_aspect': 'GENERAL_MEDICAL', 'med_key': matched_med_key}

    # 3. Check for Profit & Margin Analytics
    is_profit_query = bool(re.search(r'\b(profit|margin|margins|munafa|kamai|profitable|profitability|markup|gain|gains|highest profit|most profitable|gross profit)\b', clean, flags=re.IGNORECASE))
    if is_profit_query:
        return {'intent': 'PROFIT_MARGIN', 'sub_aspect': None, 'med_key': matched_med_key}

    # 4. Check for Patient Refill Alerts
    is_refill_query = bool(re.search(r'\b(refill|refills|repeat patient|repeat customer|dawai khatam|course khatam|due refill|refill alert|refill reminder|refill prediction|chronic patient|refill due)\b', clean, flags=re.IGNORECASE))
    if is_refill_query:
        return {'intent': 'PATIENT_REFILL', 'sub_aspect': None, 'med_key': matched_med_key}

    # 5. Check for Customer inquiries
    is_customer_query = bool(re.search(r'\b(customer|customers|client|clients|patient|patients|buyer|buyers|who bought|purchase history|spent by|customer details|customer info|phone number|mobile number|contact number|kis customer|sabse bada customer|sabse zyada kharidi|sabse jyaada kharidi|sabse jyada kharidi|kharidari|highest buyer|top buyer|top spending)\b', clean, flags=re.IGNORECASE)) or ('customer' in clean and any(w in clean for w in ['kharid', 'kharidi', 'kharidari', 'spent', 'spending', 'bada', 'zyada', 'jyaada', 'jyada']))
    if is_customer_query:
        return {'intent': 'CUSTOMER_INFO', 'sub_aspect': None, 'med_key': matched_med_key}

    # 6. Check for Demand Forecasting & Sales Velocity
    is_demand_query = bool(re.search(r'\b(demand|forecast|forecasting|prediction|predict|fast moving|fast selling|slow moving|top selling|velocity|run out|kab khatam hoga|kab khatam|reorder prediction|reorder recommendation|consumption rate|stock life)\b', clean, flags=re.IGNORECASE))
    if is_demand_query:
        return {'intent': 'DEMAND_FORECAST', 'sub_aspect': None, 'med_key': matched_med_key}

    # 7. Check for inventory queries
    if any(w in clean for w in ['stock of', 'available stock', 'how many', 'in stock', 'low stock', 'out of stock', 'units remaining', 'inventory count', 'check stock', 'show stock', 'total stock']) or (clean.startswith('stock') and not any(w in clean for w in ['price', 'expiry', 'use', 'side effect', 'demand', 'forecast', 'profit', 'margin'])):
        return {'intent': 'STOCK', 'sub_aspect': None, 'med_key': matched_med_key}

    if any(w in clean for w in ['price of', 'cost of', 'rate of', 'how much is', 'selling price', 'mrp of', 'charges of', 'pricing of']) or (clean.startswith('price') and not any(w in clean for w in ['stock', 'expiry', 'use', 'side effect', 'demand', 'forecast', 'profit', 'margin'])):
        return {'intent': 'PRICE', 'sub_aspect': None, 'med_key': matched_med_key}

    if any(w in clean for w in ['expiring soon', 'expiry date', 'expiry of', 'when does it expire', 'shelf life', 'expired batches', 'expiring in', 'validity']) or (clean.startswith('expiry') or clean.startswith('expire')):
        return {'intent': 'EXPIRY', 'sub_aspect': None, 'med_key': matched_med_key}

    if any(w in clean for w in ['today sales', 'todays sales', 'today\'s sales', 'revenue this month', 'monthly revenue', 'daily sales', 'sales summary', 'total revenue', 'invoices today']):
        return {'intent': 'SALES_REVENUE', 'sub_aspect': None, 'med_key': matched_med_key}

    if any(w in clean for w in ['supplier', 'suppliers', 'vendor', 'distributor']):
        return {'intent': 'SUPPLIERS', 'sub_aspect': None, 'med_key': matched_med_key}

    if any(w in clean for w in ['category', 'categories', 'medicine type']):
        return {'intent': 'CATEGORIES', 'sub_aspect': None, 'med_key': matched_med_key}

    if any(w in clean for w in ['help', 'command', 'commands', 'hello', 'hi', 'hey']):
        return {'intent': 'HELP', 'sub_aspect': None, 'med_key': matched_med_key}

    if matched_med_key:
        return {'intent': 'STOCK_PRICE_DETAIL', 'sub_aspect': None, 'med_key': matched_med_key}

    return {'intent': 'GENERAL_QUERY', 'sub_aspect': None, 'med_key': None}


def _format_medical_info_reply(med_data, sub_aspect, raw_prompt=""):
    """Format a detailed clinical response tailored specifically to the asked sub-aspect."""
    name = med_data['generic_name']
    drug_class = med_data['drug_class']
    disclaimer = "\n\n⚠️ *Safety Disclaimer: This clinical summary is for informational guidance only. Always consult a licensed doctor or pharmacist before taking, changing, or stopping any medication.*"

    if sub_aspect == 'SIDE_EFFECTS':
        return (
            f"### ⚠️ Side Effects & Disadvantages: {name}\n\n"
            f"• **Drug Class:** {drug_class}\n\n"
            f"**Disadvantages & Adverse Effects:**\n"
            f"{med_data['side_effects']}\n\n"
            f"**Important Precautions:**\n"
            f"{med_data['precautions']}"
            f"{disclaimer}"
        )
    elif sub_aspect == 'USES':
        return (
            f"### 💊 Medical Uses & Indications: {name}\n\n"
            f"• **Drug Class:** {drug_class}\n\n"
            f"**Therapeutic Uses:**\n"
            f"{med_data['uses']}\n\n"
            f"**Standard Dosage Guidance:**\n"
            f"{med_data['dosage']}"
            f"{disclaimer}"
        )
    elif sub_aspect == 'DOSAGE':
        return (
            f"### 📋 Dosage & Administration Guidelines: {name}\n\n"
            f"• **Drug Class:** {drug_class}\n\n"
            f"**Dosage Instructions:**\n"
            f"{med_data['dosage']}\n\n"
            f"**Safety Precautions:**\n"
            f"{med_data['precautions']}"
            f"{disclaimer}"
        )
    elif sub_aspect == 'PREGNANCY':
        return (
            f"### 🤰 Pregnancy & Lactation Safety: {name}\n\n"
            f"• **Drug Class:** {drug_class}\n\n"
            f"**Pregnancy & Breastfeeding Guidance:**\n"
            f"{med_data['pregnancy']}\n\n"
            f"**Clinical Recommendations:**\n"
            f"• Always use the minimum effective dose for the shortest duration necessary under direct obstetric supervision.\n"
            f"• Consult your healthcare provider prior to taking any medication while pregnant or nursing."
            f"{disclaimer}"
        )
    elif sub_aspect == 'PRECAUTIONS':
        return (
            f"### 🛡️ Warnings, Precautions & Interactions: {name}\n\n"
            f"• **Drug Class:** {drug_class}\n\n"
            f"**Precautions & Contraindications:**\n"
            f"{med_data['precautions']}\n\n"
            f"**Potential Side Effects:**\n"
            f"{med_data['side_effects']}"
            f"{disclaimer}"
        )
    else:  # GENERAL_MEDICAL
        return (
            f"### ℹ️ Medical Information: {name}\n\n"
            f"• **Drug Class:** {drug_class}\n\n"
            f"**Therapeutic Uses:**\n{med_data['uses']}\n\n"
            f"**Dosage Guidelines:**\n{med_data['dosage']}\n\n"
            f"**Side Effects & Disadvantages:**\n{med_data['side_effects']}\n\n"
            f"**Pregnancy Safety:**\n{med_data['pregnancy']}"
            f"{disclaimer}"
        )


def _build_medibot_context(request, prompt=""):
    """
    Build live inventory and sales context snapshot strictly isolated
    to the logged-in user's tenant scope (or all for admin).
    """
    try:
        today = timezone.now().date()
        if request and getattr(request, 'user', None) and request.user.is_authenticated:
            meds = owner_scope_queryset(
                request,
                Medicine.objects.select_related('category', 'supplier').prefetch_related('batches'),
                'created_by'
            )
            sales = owner_scope_sales(request)
            suppliers_qs = owner_scope_queryset(request, Supplier.objects.all(), 'created_by')
            categories_qs = owner_scope_queryset(request, Category.objects.all(), 'created_by')
        else:
            meds = Medicine.objects.select_related('category', 'supplier').prefetch_related('batches').all()
            sales = Sale.objects.all()
            suppliers_qs = Supplier.objects.all()
            categories_qs = Category.objects.all()

        total_meds_count = meds.count()

        # Low stock items (< 50 or below threshold)
        low_stock_items = [m for m in meds if m.is_low_stock]
        low_stock_summary = [f"{m.name} ({m.total_quantity} units left)" for m in low_stock_items[:10]]

        # Batches expiring in next 30 days
        exp_batches = Batch.objects.filter(
            medicine__in=meds,
            expiry_date__gte=today,
            expiry_date__lte=today + timedelta(days=30),
            quantity__gt=0
        ).select_related('medicine').order_by('expiry_date')
        exp_summary = [f"{b.medicine.name} (Batch: {b.batch_name}, Qty: {b.quantity}, Exp: {b.expiry_date})" for b in exp_batches[:10]]

        # Expired batches with stock
        expired_batches = Batch.objects.filter(
            medicine__in=meds,
            expiry_date__lt=today,
            quantity__gt=0
        ).select_related('medicine')
        expired_summary = [f"{b.medicine.name} (Batch: {b.batch_name}, Qty: {b.quantity}, Exp: {b.expiry_date})" for b in expired_batches[:8]]

        # Sales overview
        today_sales = sales.filter(date__date=today, status='Completed')
        today_sales_count = today_sales.count()
        today_revenue = sum((s.total_price for s in today_sales), Decimal('0.00'))

        # Catalog overview
        catalog_lines = []
        for m in meds:
            batches_info = [f"{b.batch_name}(qty:{b.quantity}, exp:{b.expiry_date})" for b in m.batches.all()]
            b_str = ', '.join(batches_info) if batches_info else 'No batches recorded'
            cat_name = m.category.name if m.category else 'General'
            sup_name = m.supplier.name if m.supplier else 'N/A'
            catalog_lines.append(
                f"- {m.name} | Category: {cat_name} | Price: Rs.{m.price} | Total Stock: {m.total_quantity} units | Batches: [{b_str}] | Supplier: {sup_name}"
            )

        context_lines = [
            f"TODAY'S DATE: {today}",
            f"TOTAL REGISTERED MEDICINES: {total_meds_count}",
            f"LOW STOCK COUNT: {len(low_stock_items)} items" + (f" -> {', '.join(low_stock_summary)}" if low_stock_summary else " (None)"),
            f"EXPIRING WITHIN 30 DAYS: {exp_batches.count()} batches" + (f" -> {', '.join(exp_summary)}" if exp_summary else " (None)"),
            f"EXPIRED BATCHES (WITH STOCK): {expired_batches.count()} batches" + (f" -> {', '.join(expired_summary)}" if expired_summary else " (None)"),
            f"CATEGORIES: {', '.join(list(categories_qs.values_list('name', flat=True))) or 'None'}",
            f"SUPPLIERS: {', '.join(list(suppliers_qs.values_list('name', flat=True))) or 'None'}",
            f"TODAY'S COMPLETED SALES: {today_sales_count} bills, Total Revenue: Rs.{today_revenue:.2f}",
            "\nLIVE MEDICINE INVENTORY CATALOG:",
            "\n".join(catalog_lines) if catalog_lines else "No medicines currently registered in inventory."
        ]
        return "\n".join(context_lines)
    except Exception as exc:
        logger.warning("Error building MediBot context: %s", exc)
        return "LIVE DATABASE UNAVAILABLE"


def _fuzzy_match_medicines(clean_query, request=None):
    """
    Find medicines in the user's database matching brand aliases,
    tokens, and typo-tolerant n-grams with SequenceMatcher.
    """
    if request and getattr(request, 'user', None) and request.user.is_authenticated:
        meds_qs = owner_scope_queryset(
            request,
            Medicine.objects.select_related('category', 'supplier').prefetch_related('batches'),
            'created_by'
        )
    else:
        meds_qs = Medicine.objects.select_related('category', 'supplier').prefetch_related('batches').all()

    all_meds = list(meds_qs)
    if not all_meds:
        return []

    expanded_query = clean_query
    for brand_k, generic_v in MEDIBOT_BRAND_ALIASES.items():
        if re.search(r'\b' + re.escape(brand_k) + r'\b', clean_query, flags=re.IGNORECASE):
            expanded_query += f" {generic_v.lower()}"

    STOP_WORDS = {
        'what', 'is', 'the', 'of', 'and', 'for', 'in', 'a', 'an', 'to', 'how', 'much',
        'many', 'do', 'we', 'have', 'you', 'there', 'check', 'show', 'tell', 'me', 'any',
        'please', 'can', 'find', 'search', 'get', 'give', 'detail', 'details', 'info',
        'information', 'record', 'records', 'medicine', 'medicines', 'tablet', 'tablets',
        'syrup', 'syrups', 'capsule', 'capsules', 'mg', 'mcg', 'gm', 'g', 'iu', '%',
        'stock', 'price', 'cost', 'rate', 'expiry', 'expir', 'expired', 'expiring',
        'available', 'availability', 'quantity', 'units', 'batch', 'batches'
    }

    raw_tokens = expanded_query.split()
    tokens = [t for t in raw_tokens if t not in STOP_WORDS and len(t) >= 2]

    candidates = list(tokens)
    for i in range(len(raw_tokens) - 1):
        candidates.append(f"{raw_tokens[i]} {raw_tokens[i+1]}")
    for i in range(len(raw_tokens) - 2):
        candidates.append(f"{raw_tokens[i]} {raw_tokens[i+1]} {raw_tokens[i+2]}")

    matched_scores = {}
    for med in all_meds:
        med_lower = med.name.lower()
        med_base = med_lower.split()[0]

        if med_base in expanded_query or med_lower in expanded_query:
            matched_scores[med.id] = (1.0, med)
            continue

        best_score = 0.0
        for cand in candidates:
            r1 = difflib.SequenceMatcher(None, cand, med_base).ratio()
            r2 = difflib.SequenceMatcher(None, cand, med_lower).ratio()
            best_score = max(best_score, r1, r2)

        if best_score >= 0.65:
            matched_scores[med.id] = (best_score, med)

    sorted_matches = sorted(matched_scores.values(), key=lambda x: x[0], reverse=True)
    return [m[1] for m in sorted_matches]


def _fuzzy_match_customers(clean_query, request=None):
    """
    Find customers matching name tokens, phone numbers, or emails
    scoped to the logged-in pharmacy tenant.
    """
    if request and getattr(request, 'user', None) and request.user.is_authenticated:
        custs = owner_scope_queryset(request, Customer.objects.all(), 'created_by')
    else:
        custs = Customer.objects.all()

    # 1. Phone number match
    phone_match = re.search(r'\b\d{7,15}\b', clean_query)
    if phone_match:
        phone_results = list(custs.filter(contact_number__icontains=phone_match.group(0)))
        if phone_results:
            return phone_results

    # 2. Clean query
    clean_name = re.sub(
        r'\b(customer|customers|client|clients|patient|patients|buyer|buyers|details|detail|info|information|history|purchase|purchases|orders|order|spent|spending|of|for|about|tell|me|show|check|find|who|is|the|give|ka|ki|ke|ko|se|me|mein|par|pe|batao|bataiye|dikhao|dikhaye|chahiye|do|de\s+do|karo|karein|hai|hain|kya|records|record|mobile|phone|contact|number|sabse|zyada|jyaada|jyada|bada|bade|highest|top|best|most|kharidi|kharidari|kharida|khareeda|khareedari|shopping|ne|kis|kisne)\b',
        ' ',
        clean_query,
        flags=re.IGNORECASE
    )
    clean_name = re.sub(r'[^\w\s]', ' ', clean_name)
    clean_name = re.sub(r'\s+', ' ', clean_name).strip()

    if not clean_name or len(clean_name) < 2:
        return []

    # Direct icontains match
    direct = list(custs.filter(name__icontains=clean_name))
    if direct:
        return direct

    # Token match (e.g. if multi-word name, check if any token matches)
    tokens = [t for t in clean_name.split() if len(t) >= 3]
    for t in tokens:
        t_matches = list(custs.filter(name__icontains=t))
        if t_matches:
            return t_matches

    # SequenceMatcher fuzzy match
    matched = []
    for c in custs:
        c_lower = c.name.lower()
        if clean_name in c_lower or c_lower in clean_name:
            matched.append((1.0, c))
            continue
        ratio = difflib.SequenceMatcher(None, clean_name, c_lower).ratio()
        if ratio >= 0.6:
            matched.append((ratio, c))

    matched.sort(key=lambda x: x[0], reverse=True)
    return [m[1] for m in matched]


def _fuzzy_match_suppliers(clean_query, request=None):
    """Find suppliers matching name tokens scoped to user."""
    if request and getattr(request, 'user', None) and request.user.is_authenticated:
        sups = owner_scope_queryset(request, Supplier.objects.all(), 'created_by')
    else:
        sups = Supplier.objects.all()

    clean_name = re.sub(r'\b(supplier|suppliers|vendor|vendors|distributor|details|detail|info|information|contact|phone|ka|ki|ke|batao|do|dikhao|records)\b', ' ', clean_query, flags=re.IGNORECASE)
    clean_name = re.sub(r'\s+', ' ', clean_name).strip()
    if not clean_name or len(clean_name) < 2:
        return []
    return list(sups.filter(name__icontains=clean_name))


def _format_customer_card(customers, scoped_sales):
    """Format structured, beautiful customer card with purchase history and action buttons."""
    import urllib.parse
    lines = []
    for c in customers[:3]:
        c_sales = scoped_sales.filter(customer=c, status='Completed').order_by('-date')
        total_spent = sum((s.total_price for s in c_sales), Decimal('0.00'))
        total_orders = c_sales.count()
        last_sale = c_sales.first()
        if last_sale and last_sale.items.exists():
            items_str = ", ".join([f"{it.medicine.name} (x{it.quantity})" for it in last_sale.items.all()[:3]])
            last_order_str = f"{last_sale.date.strftime('%Y-%m-%d')} — {items_str} (Rs.{last_sale.total_price})"
        else:
            last_order_str = "No recent purchases"
        membership_badge = "⭐ Permanent Member" if c.is_permanent else "Standard Customer"
        wa_msg = f"Hello {c.name}, greeting from PharmaCare Pharmacy! How can we assist you with your medicine prescription today?"
        wa_btn = f"[[WA:{c.contact_number}|{wa_msg}|WhatsApp {c.name.split()[0]}]]"
        lines.append(
            f"### 👤 Customer: {c.name}\n"
            f"• **Contact Number:** `{c.contact_number}`\n"
            f"• **Email:** {c.email or 'N/A'}\n"
            f"• **Membership Status:** {membership_badge}\n"
            f"• **Total Completed Orders:** **{total_orders} bills**\n"
            f"• **Lifetime Spending:** **Rs.{total_spent:.2f}**\n"
            f"• **Last Purchase:** {last_order_str}\n"
            f"• **Quick Action:** {wa_btn}"
        )
    return "\n\n".join(lines)


def _format_markdown_to_html(text):
    """Format markdown text into clean, styled HTML for the chat window with action buttons."""
    if not text:
        return ""

    out = html.escape(text.strip())

    # Headers (###, ##, #)
    out = re.sub(
        r'(?m)^###\s+(.*?)$',
        r'<div style="font-weight:700; font-size:13.5px; color:#0891b2; margin-top:8px; margin-bottom:3px;">\1</div>',
        out
    )
    out = re.sub(
        r'(?m)^##\s+(.*?)$',
        r'<div style="font-weight:700; font-size:14.5px; color:#0e7490; margin-top:9px; margin-bottom:4px;">\1</div>',
        out
    )
    out = re.sub(
        r'(?m)^#\s+(.*?)$',
        r'<div style="font-weight:800; font-size:15.5px; color:#0e7490; margin-top:10px; margin-bottom:5px;">\1</div>',
        out
    )

    # Bold **text** -> <strong>text</strong>
    out = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', out)

    # Inline code `code`
    out = re.sub(r'`([^`]+)`', r'<code style="background:rgba(8,145,178,0.12); color:#0891b2; padding:1px 5px; border-radius:4px; font-size:12px; font-family:monospace;">\1</code>', out)

    # Action tags [[WA:phone|text|label]], [[PO:med_id|label]], [[LINK:url|label]]
    def _render_wa(m):
        phone, msg, label = m.group(1), m.group(2), m.group(3)
        clean_phone = re.sub(r'[^\d]', '', phone)
        if len(clean_phone) == 10:
            clean_phone = '91' + clean_phone
        import urllib.parse
        encoded = urllib.parse.quote(html.unescape(msg))
        return f'<a href="https://wa.me/{clean_phone}?text={encoded}" target="_blank" class="medibot-action-btn wa-btn" style="display:inline-flex; align-items:center; gap:4px; padding:3px 9px; background:#25D366; color:#ffffff; border-radius:6px; font-size:11.5px; font-weight:600; text-decoration:none; margin:3px 4px 3px 0; box-shadow:0 1px 3px rgba(37,211,102,0.3);">📲 {label}</a>'

    def _render_po(m):
        med_id, label = m.group(1), m.group(2)
        return f'<a href="/purchase-orders/create/?medicine={med_id}" class="medibot-action-btn po-btn" style="display:inline-flex; align-items:center; gap:4px; padding:3px 9px; background:#0891b2; color:#ffffff; border-radius:6px; font-size:11.5px; font-weight:600; text-decoration:none; margin:3px 4px 3px 0; box-shadow:0 1px 3px rgba(8,145,178,0.3);">📥 {label}</a>'

    def _render_link(m):
        url, label = m.group(1), m.group(2)
        return f'<a href="{url}" class="medibot-action-btn link-btn" style="display:inline-flex; align-items:center; gap:4px; padding:3px 9px; background:#0284c7; color:#ffffff; border-radius:6px; font-size:11.5px; font-weight:600; text-decoration:none; margin:3px 4px 3px 0;">{label}</a>'

    out = re.sub(r'\[\[WA:([^|]+)\|([^|]+)\|([^\]]+)\]\]', _render_wa, out)
    out = re.sub(r'\[\[PO:([^|]+)\|([^\]]+)\]\]', _render_po, out)
    out = re.sub(r'\[\[LINK:([^|]+)\|([^\]]+)\]\]', _render_link, out)

    # Bullet points (* or -)
    out = re.sub(
        r'(?m)^[\*\-]\s+(.*?)$',
        r'<div style="display:flex; align-items:flex-start; margin-left:4px; margin-bottom:3px;"><span style="color:#0891b2; margin-right:6px; font-weight:bold;">•</span><span>\1</span></div>',
        out
    )

    # Numbered lists
    out = re.sub(
        r'(?m)^(\d+)\.\s+(.*?)$',
        r'<div style="display:flex; align-items:flex-start; margin-left:4px; margin-bottom:3px;"><span style="font-weight:600; color:#0891b2; margin-right:6px;">\1.</span><span>\2</span></div>',
        out
    )

    # Paragraph breaks & linebreaks
    out = out.replace('\n\n', '<div style="height:6px;"></div>').replace('\n', '<br>')
    return out


def _call_gemini_api(prompt, context_str=None, intent_info=None, request=None):
    """
    Call Gemini Generative Language API securely with intent-specific system framing.
    Returns response text string or None on failure/timeout.
    """
    import os
    api_key = (
        getattr(django_settings, 'GEMINI_API_KEY', '') or
        os.environ.get('GEMINI_API_KEY', '') or
        ''
    ).strip()

    if not api_key:
        return None

    if intent_info is None:
        intent_info = _classify_query_intent(prompt)

    if context_str is None:
        context_str = _build_medibot_context(request, prompt) if intent_info.get('intent') != 'MEDICAL_INFO' else ""

    configured_model = getattr(django_settings, 'GEMINI_MODEL', 'gemini-flash-lite-latest') or 'gemini-flash-lite-latest'
    intent = intent_info.get('intent') if intent_info else 'GENERAL_QUERY'
    sub_aspect = intent_info.get('sub_aspect') if intent_info else None

    # Intent-specific prompt framing
    if intent == 'MEDICAL_INFO':
        system_instruction_text = (
            "You are MediBot, an expert clinical pharmacy assistant for PharmaCare.\n"
            f"The user is asking a medical/clinical question regarding: '{prompt}'.\n"
            f"Identified aspect: {sub_aspect or 'clinical information'}.\n\n"
            "CRITICAL INSTRUCTIONS:\n"
            "1. Focus STRICTLY on pharmaceutical and clinical guidance (uses, indications, side effects, disadvantages, dosage guidelines, safety in pregnancy/breastfeeding, contraindications, and drug interactions).\n"
            "2. DO NOT output pharmacy stock quantities, unit counts, warehouse batches, or supplier names for medical questions.\n"
            "3. Answer thoroughly, directly, and politely using clean markdown headings and bullet points.\n"
            "4. Always conclude with a medical safety disclaimer: '⚠️ *Safety Disclaimer: This clinical summary is for informational guidance only. Always consult a licensed physician or pharmacist for medical advice.*'"
        )
    elif intent in ['STOCK', 'PRICE', 'EXPIRY', 'SALES_REVENUE', 'SUPPLIERS', 'CATEGORIES', 'STOCK_PRICE_DETAIL']:
        system_instruction_text = (
            "You are MediBot, a pharmacy inventory assistant for PharmaCare.\n"
            "Answer the user's question accurately using the live pharmacy database context provided below.\n\n"
            "=== LIVE PHARMACY DATABASE CONTEXT ===\n"
            f"{context_str}\n"
            "======================================\n\n"
            "INSTRUCTIONS:\n"
            "1. Use the exact numbers (units, prices, batches, expiry dates, revenue) from the context.\n"
            "2. Format response cleanly with markdown highlights and bullet points."
        )
    else:
        system_instruction_text = (
            "You are MediBot, a helpful pharmacy and health assistant for PharmaCare.\n"
            "Respond politely and concisely to the user's message using markdown formatting.\n\n"
            "=== PHARMACY CONTEXT ===\n"
            f"{context_str}\n"
            "========================"
        )

    payload = {
        'systemInstruction': {
            'parts': [{'text': system_instruction_text}]
        },
        'contents': [
            {'role': 'user', 'parts': [{'text': prompt}]}
        ],
        'generationConfig': {
            'temperature': 0.2,
            'maxOutputTokens': 800,
        },
    }

    payload_bytes = json.dumps(payload).encode('utf-8')
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{configured_model}:generateContent"
    req = urllib.request.Request(
        url,
        data=payload_bytes,
        headers={
            'Content-Type': 'application/json',
            'x-goog-api-key': api_key
        },
        method='POST'
    )

    try:
        with urllib.request.urlopen(req, timeout=2.0) as response:
            res_json = json.loads(response.read().decode('utf-8'))
            candidates = res_json.get('candidates') or []
            if candidates:
                parts = candidates[0].get('content', {}).get('parts', []) or []
                reply_text = ''.join(p.get('text', '') for p in parts if isinstance(p, dict)).strip()
                if reply_text:
                    return reply_text
    except Exception as exc:
        logger.warning("Gemini AI API note: %s", exc)
        return None

    return None


_call_gemini_ai = _call_gemini_api


def _medibot_local_engine(prompt, request=None):
    """
    High-reliability, instant offline database & medical engine with precise
    intent classification, clinical knowledge base, stock/pricing lookups, expiry alerts, and sales summaries.
    """
    raw_query = (prompt or '').strip()
    clean_query = _clean_user_prompt(raw_query)
    today = timezone.now().date()

    if not clean_query:
        return "👋 Hi! How can I help you today? You can ask about medicine stock, prices, expiry alerts, today's sales, or clinical medicine info."

    session = getattr(request, 'session', None) if request else None
    session_med_key = session.get('medibot_last_med_key') if session else None
    if session_med_key and bool(re.search(r'\b(iska|iski|isko|it|its|this medicine|this drug|this item|inka|inki)\b', clean_query, flags=re.IGNORECASE)):
        clean_query = f"{clean_query} {session_med_key}"

    # Classify intent and sub-aspect
    intent_data = _classify_query_intent(clean_query, session=session)
    intent = intent_data['intent']
    sub_aspect = intent_data['sub_aspect']
    med_key = intent_data['med_key']

    # 1. MEDICAL_INFO INTENT -> Never return inventory cards!
    if intent == 'MEDICAL_INFO':
        target_med_key = med_key
        if not target_med_key:
            # Try fuzzy match against medical knowledge keys
            for k in MEDIBOT_MEDICAL_KNOWLEDGE.keys():
                if k in clean_query:
                    target_med_key = k
                    break
        if not target_med_key:
            # Check brand aliases
            for b_k, g_v in MEDIBOT_BRAND_ALIASES.items():
                if b_k in clean_query:
                    target_med_key = g_v.lower()
                    break

        if target_med_key and target_med_key in MEDIBOT_MEDICAL_KNOWLEDGE:
            med_info = MEDIBOT_MEDICAL_KNOWLEDGE[target_med_key]
            return _format_medical_info_reply(med_info, sub_aspect, raw_query)

        # If medicine not in hardcoded medical knowledge dictionary, provide general medical guidance
        return (
            f"### ℹ️ Medical Information Inquiry: '{raw_query}'\n\n"
            f"• MediBot recognizes this as a clinical/medical question.\n"
            f"• When taking prescription or over-the-counter medications, always adhere strictly to the dosage and frequency prescribed by your doctor.\n"
            f"• Be aware of potential adverse reactions, contraindications with pre-existing conditions, and drug interactions.\n\n"
            f"⚠️ *Safety Disclaimer: Always consult a licensed doctor or pharmacist for personalized clinical guidance and verified prescription advice.*"
        )

    # Scoped database queries
    if request and getattr(request, 'user', None) and request.user.is_authenticated:
        scoped_meds = owner_scope_queryset(request, Medicine.objects.select_related('category', 'supplier').prefetch_related('batches'), 'created_by')
        scoped_sales = owner_scope_sales(request)
        scoped_suppliers = owner_scope_queryset(request, Supplier.objects.all(), 'created_by')
        scoped_categories = owner_scope_queryset(request, Category.objects.all(), 'created_by')
        scoped_customers = owner_scope_queryset(request, Customer.objects.all(), 'created_by')
    else:
        scoped_meds = Medicine.objects.select_related('category', 'supplier').prefetch_related('batches').all()
        scoped_sales = Sale.objects.all()
        scoped_suppliers = Supplier.objects.all()
        scoped_categories = Category.objects.all()
        scoped_customers = Customer.objects.all()

    # 2. CUSTOMER_INFO INTENT
    if intent == 'CUSTOMER_INFO':
        # Check if ranking / highest buyer query (in English or Hinglish)
        is_top_spending_query = bool(re.search(
            r'\b(sabse\s+(zyada|jyaada|jyada|bada|bade)\s+(kharid|kharidi|kharidari|khareeda|khareedari|spending|spent|orders|shopping|customer|customers|buyer|buyers)|highest\s+(spending|buyer|buyers|purchases|orders)|top\s+(spending|customers|customer|buyers|buyer)|best\s+(customer|customers)|max\s+spending|maximum\s+spending|highest\s+buyer)\b',
            clean_query,
            flags=re.IGNORECASE
        )) or ('sabse' in clean_query and any(w in clean_query for w in ['kharid', 'kharidi', 'kharidari', 'khareeda', 'shopping', 'spent', 'zyada', 'jyaada', 'jyada', 'bada', 'bade'])) or (clean_query.startswith('customer') and any(w in clean_query for w in ['sabse', 'highest', 'top', 'best', 'kharid', 'spending', 'spent']))

        if is_top_spending_query:
            cust_spending = []
            for c in scoped_customers:
                c_sales = scoped_sales.filter(customer=c, status='Completed')
                spent = sum((s.total_price for s in c_sales), Decimal('0.00'))
                order_count = c_sales.count()
                if order_count > 0:
                    cust_spending.append({
                        'customer': c,
                        'total_spent': spent,
                        'order_count': order_count,
                    })

            if cust_spending:
                cust_spending.sort(key=lambda x: (x['total_spent'], x['order_count']), reverse=True)
                lines = []
                for idx, item in enumerate(cust_spending[:5], 1):
                    c = item['customer']
                    wa_msg = f"Hello {c.name}, greeting from PharmaCare! Thank you for being our valued customer. How can we assist you with your health and medicine needs today?"
                    wa_btn = f"[[WA:{c.contact_number}|{wa_msg}|WhatsApp {c.name.split()[0]}]]"
                    badge = "⭐ Top Spender" if idx == 1 else f"#{idx} Buyer"
                    lines.append(
                        f"• **{idx}. {c.name}** (`{c.contact_number}`) — Total Spent: **Rs.{item['total_spent']:.2f}** ({item['order_count']} bills) | {badge} {wa_btn}"
                    )
                return (
                    f"🏆 **Top Spending Customers (Highest Buyers):**\n\n" +
                    "\n".join(lines) +
                    "\n\n💡 *Tip: Click WhatsApp buttons to directly message top customers.*"
                )
            else:
                return "👤 **Customer Spending:** Abhi tak kisi registered customer ke completed orders record nahi hue hain."

        matched_customers = _fuzzy_match_customers(raw_query, request=request)
        if matched_customers:
            return _format_customer_card(matched_customers, scoped_sales)

        # Permanent / Loyal customer list
        if any(w in clean_query for w in ['permanent', 'permanent customer', 'permanent customers', 'loyal member', 'loyal members', 'vip', 'permanent member', 'permanent members', 'permanent bala', 'permanent wale', 'sign bana', 'permanent sign']):
            perm_custs = list(scoped_customers.filter(is_permanent=True))
            if not perm_custs:
                return "⭐ No permanent members are currently marked in your pharmacy records."
            lines = []
            for c in perm_custs:
                c_sales = scoped_sales.filter(customer=c, status='Completed')
                spent = sum((s.total_price for s in c_sales), Decimal('0.00'))
                wa_msg = f"Hello {c.name}, greeting from PharmaCare! As our valued permanent member, how can we assist you today?"
                wa_btn = f"[[WA:{c.contact_number}|{wa_msg}|WhatsApp {c.name.split()[0]}]]"
                lines.append(f"• **{c.name}** (`{c.contact_number}`) — Orders: **{c_sales.count()} bills**, Total Spent: **Rs.{spent:.2f}** {wa_btn}")
            perm_btn = "[[LINK:/customer/|👥 Open Customer Management]]"
            return f"⭐ **Permanent Member Customers ({len(perm_custs)} total):**\n\n" + "\n".join(lines) + f"\n\n{perm_btn}"

        # General customer list (All customers)
        if any(w in clean_query for w in ['top', 'list', 'all', 'spending', 'highest', 'best', 'loyal', 'registered', 'customers', 'pure', 'saare', 'sare', 'batao', 'dikhaye']):
            all_custs = list(scoped_customers)
            if not all_custs:
                return "👤 No customers are currently registered in your pharmacy account."
            lines = []
            for idx, c in enumerate(all_custs, 1):
                c_sales = scoped_sales.filter(customer=c, status='Completed')
                spent = sum((s.total_price for s in c_sales), Decimal('0.00'))
                perm_badge = " ⭐" if c.is_permanent else ""
                lines.append(f"• **{idx}. {c.name}**{perm_badge} (`{c.contact_number}`) — Orders: **{c_sales.count()}**, Spent: **Rs.{spent:.2f}**")
            total_cust_count = len(all_custs)
            crm_btn = "[[LINK:/customer/|👥 Open Full Customer Management (96 total)]]"
            return f"👤 **All Registered Customers ({total_cust_count} total):**\n\n" + "\n".join(lines) + f"\n\n{crm_btn}"

        return f"👤 I could not find customer records matching **'{raw_query}'**. Try searching by customer name (e.g. *'Customer Rahul'*) or phone number (*'Customer 9876543210'*)."

    # 3. DEMAND_FORECAST INTENT
    if intent == 'DEMAND_FORECAST':
        start_30d = today - timedelta(days=30)
        sales_30d = scoped_sales.filter(date__date__gte=start_30d, date__date__lte=today, status='Completed')
        items_30d = SaleItem.objects.filter(sale__in=sales_30d)

        matching_meds = _fuzzy_match_medicines(clean_query, request=request)
        if matching_meds:
            lines = []
            for med in matching_meds[:3]:
                med_items = items_30d.filter(medicine=med)
                units_sold = med_items.aggregate(t=Sum('quantity'))['t'] or 0
                daily_velocity = float(units_sold) / 30.0
                curr_stock = med.total_quantity

                if daily_velocity > 0:
                    days_left = int(curr_stock / daily_velocity)
                    runout_date = today + timedelta(days=days_left)
                    if daily_velocity >= 5:
                        demand_badge = "🔥 High Demand (Fast-Moving)"
                    elif daily_velocity >= 1:
                        demand_badge = "🟢 Moderate Demand"
                    else:
                        demand_badge = "🟡 Low Demand (Slow-Moving)"

                    reorder_target = int(daily_velocity * 30)
                    reorder_qty = max(0, reorder_target - curr_stock)
                    if reorder_qty > 0:
                        reorder_tip = f"Recommended reorder of **{reorder_qty} units** to maintain a 30-day stock buffer."
                    else:
                        reorder_tip = "Current stock is healthy and sufficient for the next 30 days."

                    lines.append(
                        f"### 📈 Demand Forecast: {med.name}\n"
                        f"• **Current Available Stock:** **{curr_stock} units**\n"
                        f"• **30-Day Sales Volume:** **{units_sold} units** sold\n"
                        f"• **Daily Velocity:** **~{daily_velocity:.1f} units/day**\n"
                        f"• **Demand Level:** {demand_badge}\n"
                        f"• **Estimated Stock Runway:** **~{days_left} days** (Est. stock-out: {runout_date.strftime('%Y-%m-%d')})\n"
                        f"• **Reorder Recommendation:** {reorder_tip}"
                    )
                else:
                    lines.append(
                        f"### 📈 Demand Forecast: {med.name}\n"
                        f"• **Current Available Stock:** **{curr_stock} units**\n"
                        f"• **30-Day Sales Volume:** **0 units** (No sales in past 30 days)\n"
                        f"• **Demand Level:** ⚪ Dormant / Zero Demand\n"
                        f"• **Reorder Recommendation:** Current stock is sufficient; no immediate reorder needed."
                    )
            return "\n\n".join(lines)

        # General demand forecast leaderboard
        top_items = items_30d.values('medicine__name', 'medicine__id').annotate(total_qty=Sum('quantity')).order_by('-total_qty')[:5]
        if top_items:
            lines = []
            for item in top_items:
                med_obj = scoped_meds.filter(id=item['medicine__id']).first() if hasattr(scoped_meds, 'filter') else next((m for m in scoped_meds if m.id == item['medicine__id']), None)
                curr_stock = med_obj.total_quantity if med_obj else 0
                units = item['total_qty'] or 0
                vel = float(units) / 30.0
                days_left = int(curr_stock / vel) if vel > 0 else 999
                urgency = " ⚠️ Low Runway!" if days_left <= 10 else ""
                lines.append(f"• **{item['medicine__name']}**: Sold **{units} units** (~{vel:.1f}/day) | Stock: **{curr_stock}** (~{days_left} days left){urgency}")
            return (
                f"📈 **Top Fast-Moving Medicines (30-Day Demand Forecast):**\n\n"
                + "\n".join(lines) +
                "\n\n💡 *Tip: Ask 'demand forecast of [medicine name]' for detailed stock run-out projections and reorder calculations.*"
            )

        return "📈 **Demand Forecasting:** No sales recorded in the past 30 days to generate forecast predictions. As sales happen, MediBot will automatically predict demand velocities and stock-out timelines."

    # 4. PROFIT_MARGIN INTENT
    if intent == 'PROFIT_MARGIN':
        matching_meds = _fuzzy_match_medicines(clean_query, request=request)
        if matching_meds:
            lines = []
            for med in matching_meds[:3]:
                batches = med.batches.filter(quantity__gt=0)
                if not batches.exists():
                    batches = med.batches.all()

                batches_with_cost = [b for b in batches if b.purchase_price and b.purchase_price > 0]
                if batches_with_cost:
                    latest_b = sorted(batches_with_cost, key=lambda b: b.id, reverse=True)[0]
                    cost_price = latest_b.purchase_price
                else:
                    cost_price = Decimal('0.00')

                selling_price = med.price or Decimal('0.00')
                gross_profit = selling_price - cost_price if cost_price > 0 else Decimal('0.00')
                margin_pct = ((gross_profit / selling_price) * Decimal('100.0')) if (selling_price > 0 and cost_price > 0) else Decimal('0.0')

                start_30d = today - timedelta(days=30)
                sales_30d = scoped_sales.filter(date__date__gte=start_30d, date__date__lte=today, status='Completed')
                med_items = SaleItem.objects.filter(sale__in=sales_30d, medicine=med)
                total_units_sold = med_items.aggregate(t=Sum('quantity'))['t'] or 0
                med_30d_rev = sum((it.price * it.quantity for it in med_items), Decimal('0.00'))
                med_30d_profit = sum(((it.price - (it.batch.purchase_price if it.batch and it.batch.purchase_price > 0 else cost_price)) * it.quantity for it in med_items), Decimal('0.00')) if cost_price > 0 else Decimal('0.00')

                if margin_pct >= 40:
                    badge = "💎 Ultra-High Margin"
                elif margin_pct >= 25:
                    badge = "🟢 High Margin"
                elif margin_pct >= 10:
                    badge = "🟡 Moderate Margin"
                else:
                    badge = "⚪ Low Margin / Standard"

                po_btn = f"[[PO:{med.id}|Order {med.name}]]"
                cost_display = f"Rs.{cost_price:.2f}" if cost_price > 0 else "N/A (No batch cost)"
                profit_display = f"Rs.{gross_profit:.2f}" if cost_price > 0 else "N/A"
                margin_display = f"{margin_pct:.1f}%" if cost_price > 0 else "N/A"

                lines.append(
                    f"### 💵 Profit & Margin Analysis: {med.name}\n"
                    f"• **Selling Price (MRP):** **Rs.{selling_price:.2f}**\n"
                    f"• **Estimated Cost Price:** **{cost_display}**\n"
                    f"• **Gross Profit / Unit:** **{profit_display}** ({badge})\n"
                    f"• **Profit Margin:** **{margin_display}**\n"
                    f"• **30-Day Sales Volume:** **{total_units_sold} units** (Revenue: Rs.{med_30d_rev:.2f})\n"
                    f"• **30-Day Gross Profit Earned:** **Rs.{med_30d_profit:.2f}**\n"
                    f"• **Restock Action:** {po_btn}"
                )
            return "\n\n".join(lines)

        # General profit ranking / leaderboard across all medicines
        start_30d = today - timedelta(days=30)
        sales_30d = scoped_sales.filter(date__date__gte=start_30d, date__date__lte=today, status='Completed')
        items_30d = list(SaleItem.objects.filter(sale__in=sales_30d).select_related('medicine', 'batch'))

        med_profit_rank = []
        for med in scoped_meds:
            batches = med.batches.filter(purchase_price__gt=0)
            if batches.exists():
                latest_b = batches.order_by('-id').first()
                cost = latest_b.purchase_price
                selling = med.price or Decimal('0.00')
                unit_profit = selling - cost
                margin = (unit_profit / selling * 100) if selling > 0 else Decimal('0.0')

                m_items = [it for it in items_30d if it.medicine_id == med.id]
                units_sold = sum((it.quantity for it in m_items), 0)
                tot_profit = sum(((it.price - (it.batch.purchase_price if it.batch and it.batch.purchase_price > 0 else cost)) * it.quantity for it in m_items), Decimal('0.00'))

                med_profit_rank.append({
                    'med': med,
                    'cost': cost,
                    'selling': selling,
                    'unit_profit': unit_profit,
                    'margin': margin,
                    'units_sold': units_sold,
                    'tot_profit': tot_profit
                })

        if med_profit_rank:
            med_profit_rank.sort(key=lambda x: (x['tot_profit'], x['margin']), reverse=True)
            lines = []
            for item in med_profit_rank[:5]:
                m = item['med']
                po_btn = f"[[PO:{m.id}|Reorder]]"
                lines.append(
                    f"• **{m.name}**: Margin: **{item['margin']:.1f}%** (Profit: **Rs.{item['unit_profit']:.2f}/unit**) | 30d Profit: **Rs.{item['tot_profit']:.2f}** {po_btn}"
                )
            return (
                "💵 **Top High-Margin & Profitable Medicines:**\n\n" +
                "\n".join(lines) +
                "\n\n💡 *Tip: Ask 'profit of [medicine name]' for individual unit economics and reorder shortcuts.*"
            )

        return "💵 **Profit Analytics:** Please record batch purchase prices in your inventory to calculate profit margins and net earnings."

    # 5. PATIENT_REFILL INTENT
    if intent == 'PATIENT_REFILL':
        matched_custs = _fuzzy_match_customers(raw_query, request=request)
        if matched_custs:
            lines = []
            for c in matched_custs[:2]:
                c_sales = scoped_sales.filter(customer=c, status='Completed').order_by('-date')[:5]
                refill_items = []
                for sale in c_sales:
                    sale_date = sale.date.date()
                    for it in sale.items.select_related('medicine').all():
                        if not it.medicine:
                            continue
                        days_supply = 30 if it.quantity >= 20 else max(7, it.quantity)
                        due_date = sale_date + timedelta(days=days_supply)
                        days_diff = (due_date - today).days
                        refill_items.append({
                            'medicine': it.medicine,
                            'last_date': sale_date,
                            'qty': it.quantity,
                            'due_date': due_date,
                            'days_diff': days_diff
                        })

                if refill_items:
                    refill_lines = []
                    for r in refill_items[:3]:
                        if r['days_diff'] < 0:
                            status_str = f"⚠️ Overdue by {abs(r['days_diff'])} days ({r['due_date'].strftime('%d %b')})"
                        elif r['days_diff'] <= 7:
                            status_str = f"🔔 Due in {r['days_diff']} days ({r['due_date'].strftime('%d %b')})"
                        else:
                            status_str = f"🟢 Supply active ({r['days_diff']} days remaining)"

                        wa_msg = f"Hello {c.name}, greeting from PharmaCare. Your refill for {r['medicine'].name} was due around {r['due_date'].strftime('%d %b')}. Would you like us to prepare your prescription?"
                        wa_btn = f"[[WA:{c.contact_number}|{wa_msg}|Refill WhatsApp]]"
                        refill_lines.append(f"  - **{r['medicine'].name}** (Qty: {r['qty']}) — {status_str} {wa_btn}")

                    lines.append(
                        f"### 🔔 Patient Refill Status: {c.name}\n"
                        f"• **Contact Number:** `{c.contact_number}`\n"
                        f"• **Refill Projections:**\n" + "\n".join(refill_lines)
                    )
                else:
                    lines.append(f"### 🔔 Patient Refill Status: {c.name}\n• No prescription purchases found for refill tracking.")
            return "\n\n".join(lines)

        # General refill prediction across regular patients
        start_60d = today - timedelta(days=60)
        recent_sales = scoped_sales.filter(date__date__gte=start_60d, status='Completed', customer__isnull=False).select_related('customer').prefetch_related('items__medicine').order_by('-date')

        refill_alerts = []
        seen_pairs = set()
        for sale in recent_sales:
            c = sale.customer
            if not c or not c.contact_number:
                continue
            sale_date = sale.date.date()
            for it in sale.items.all():
                if not it.medicine:
                    continue
                pair_key = (c.id, it.medicine.id)
                if pair_key in seen_pairs:
                    continue
                seen_pairs.add(pair_key)

                days_supply = 30 if it.quantity >= 20 else max(7, it.quantity)
                due_date = sale_date + timedelta(days=days_supply)
                days_diff = (due_date - today).days

                if -14 <= days_diff <= 7:
                    refill_alerts.append({
                        'customer': c,
                        'medicine': it.medicine,
                        'due_date': due_date,
                        'days_diff': days_diff,
                        'qty': it.quantity
                    })

        if refill_alerts:
            refill_alerts.sort(key=lambda x: x['days_diff'])
            lines = []
            for r in refill_alerts[:6]:
                c = r['customer']
                m = r['medicine']
                if r['days_diff'] < 0:
                    badge = f"⚠️ Overdue ({abs(r['days_diff'])}d ago)"
                elif r['days_diff'] == 0:
                    badge = "🔔 Due Today"
                else:
                    badge = f"⏰ Due in {r['days_diff']}d"

                wa_msg = f"Hello {c.name}, greeting from PharmaCare. Your refill for {m.name} is scheduled for {r['due_date'].strftime('%d %b')}. Let us know if you need home delivery or pharmacy pickup!"
                wa_btn = f"[[WA:{c.contact_number}|{wa_msg}|WhatsApp {c.name.split()[0]}]]"
                lines.append(f"• **{c.name}** (`{c.contact_number}`) — **{m.name}** | {badge} {wa_btn}")

            return (
                f"🔔 **Patient Prescription Refill Alerts ({len(refill_alerts)} patients due):**\n\n" +
                "\n".join(lines) +
                "\n\n💡 *Tip: Click WhatsApp buttons to directly send instant prescription refill reminders to patients.*"
            )

        return "✅ **No Patient Refills Due:** All chronic and regular patients have active medication supplies for the next 7 days."

    # 6. STOCK INTENT
    if intent == 'STOCK':
        if any(w in clean_query for w in ['low stock', 'out of stock', 'shortage', 'reorder', 'restock']):
            low_meds = [m for m in scoped_meds if m.is_low_stock]
            if not low_meds:
                return "✅ **Good news!** There are currently no medicines in low stock in your inventory."
            lines = [f"• **{m.name}**: {m.total_quantity} units remaining (Price: Rs.{m.price})" for m in low_meds[:12]]
            return f"⚠️ **Low Stock Alert ({len(low_meds)} items need attention):**\n\n" + "\n".join(lines)

        matching_meds = _fuzzy_match_medicines(clean_query, request=request)
        if matching_meds:
            lines = []
            for m in matching_meds[:5]:
                batches = m.batches.filter(quantity__gt=0).order_by('expiry_date')
                b_info = ', '.join([f"`{b.batch_name}` (Qty: {b.quantity}, Exp: {b.expiry_date})" for b in batches]) if batches.exists() else "No active batches in stock"
                stock_badge = "🟢 In Stock" if m.total_quantity >= 50 else ("🟠 Low Stock" if m.total_quantity > 0 else "🔴 Out of Stock")
                lines.append(
                    f"### 📦 {m.name} — Stock Details\n"
                    f"• **Total Available Stock:** **{m.total_quantity} units** ({stock_badge})\n"
                    f"• **Selling Price:** Rs.{m.price}\n"
                    f"• **Category:** {m.category.name if m.category else 'General'}\n"
                    f"• **Active Batches:** {b_info}"
                )
            return "\n\n".join(lines)
        return f"📦 I could not find stock records matching **'{raw_query}'** in your pharmacy inventory."

    # 7. PRICE INTENT
    if intent == 'PRICE':
        matching_meds = _fuzzy_match_medicines(clean_query, request=request)
        if matching_meds:
            lines = []
            for m in matching_meds[:5]:
                stock_badge = "🟢 In Stock" if m.total_quantity >= 50 else ("🟠 Low Stock" if m.total_quantity > 0 else "🔴 Out of Stock")
                lines.append(
                    f"### 💰 {m.name} — Pricing Details\n"
                    f"• **Selling Price:** **Rs.{m.price}**\n"
                    f"• **Availability:** {stock_badge} ({m.total_quantity} units in stock)\n"
                    f"• **Category:** {m.category.name if m.category else 'General'}\n"
                    f"• **Supplier:** {m.supplier.name if m.supplier else 'N/A'}"
                )
            return "\n\n".join(lines)
        return f"💰 I could not find pricing records matching **'{raw_query}'** in your pharmacy inventory."

    # 8. EXPIRY INTENT
    if intent == 'EXPIRY':
        is_specific_med = bool(re.search(r'\b(expiry|expiring|expiry date)\s+(of|for|about)\s+[a-zA-Z0-9]+', clean_query, flags=re.IGNORECASE))
        if is_specific_med:
            matching_meds = _fuzzy_match_medicines(clean_query, request=request)
            if matching_meds:
                lines = []
                for m in matching_meds[:4]:
                    batches = m.batches.filter(quantity__gt=0).order_by('expiry_date')
                    b_info = ', '.join([f"`{b.batch_name}` (Qty: {b.quantity}, Exp: {b.expiry_date})" for b in batches]) if batches.exists() else "No active batches in stock"
                    lines.append(
                        f"### ⏰ {m.name} — Expiry & Batch Details\n"
                        f"• **Current Stock:** {m.total_quantity} units\n"
                        f"• **Active Batches & Expiry Dates:** {b_info}"
                    )
                return "\n\n".join(lines)

        exp_batches = Batch.objects.filter(
            medicine__in=scoped_meds,
            expiry_date__gte=today,
            expiry_date__lte=today + timedelta(days=30),
            quantity__gt=0
        ).select_related('medicine').order_by('expiry_date')
        if not exp_batches.exists():
            return "✅ **Good news!** No medicine batches in your inventory are expiring within the next 30 days."
        lines = [f"• **{b.medicine.name}** (Batch: `{b.batch_name}`) — Qty: **{b.quantity}**, Exp: **{b.expiry_date}**" for b in exp_batches[:10]]
        return f"⏰ **Medicines Expiring Soon (Next 30 Days):**\n\n" + "\n".join(lines)

    # 9. SALES & REVENUE INTENT
    if intent == 'SALES_REVENUE':
        if any(w in clean_query for w in ['month revenue', 'monthly revenue', 'revenue this month', 'this month sale', 'this month sales']):
            start_of_month = today.replace(day=1)
            month_sales = scoped_sales.filter(date__date__gte=start_of_month, date__date__lte=today, status='Completed')
            m_count = month_sales.count()
            m_rev = sum((s.total_price for s in month_sales), Decimal('0.00'))
            return (
                f"💰 **Monthly Revenue Summary ({today.strftime('%B %Y')}):**\n\n"
                f"• **Completed Invoices:** {m_count}\n"
                f"• **Total Revenue Earned:** Rs.{m_rev:.2f}"
            )
        today_sales = scoped_sales.filter(date__date=today, status='Completed')
        count = today_sales.count()
        rev = sum((s.total_price for s in today_sales), Decimal('0.00'))
        return (
            f"💰 **Today's Sales Summary ({today}):**\n\n"
            f"• **Completed Invoices:** {count}\n"
            f"• **Total Revenue Earned:** Rs.{rev:.2f}"
        )

    # 10. SUPPLIERS
    if intent == 'SUPPLIERS':
        suppliers = scoped_suppliers[:8]
        if not suppliers.exists():
            return "🏢 No suppliers are currently registered in your pharmacy account."
        lines = [f"• **{s.name}** (Phone: {s.contact_number or 'N/A'}, City: {s.location or 'N/A'})" for s in suppliers]
        return f"🏢 **Registered Suppliers ({scoped_suppliers.count()} total):**\n\n" + "\n".join(lines)

    # 11. CATEGORIES
    if intent == 'CATEGORIES':
        cats = scoped_categories[:12]
        if not cats.exists():
            return "📑 No medicine categories are currently registered in your pharmacy account."
        lines = [f"• **{c.name}**" for c in cats]
        return f"📑 **Medicine Categories ({scoped_categories.count()} total):**\n\n" + "\n".join(lines)

    # 12. HELP
    if intent == 'HELP':
        return (
            "🤖 **MediBot Quick Commands & Capabilities:**\n\n"
            "• **Stock:** *'stock of Paracetamol'* or *'show low stock'*\n"
            "• **Price:** *'price of Azithromycin'*\n"
            "• **Customer Details:** *'customer Rahul'* or *'phone 9876543210'*\n"
            "• **Profit Margins:** *'profit of Paracetamol'* or *'most profitable medicines'*\n"
            "• **Patient Refills:** *'patient refill alerts'* or *'refill for Rahul'*\n"
            "• **Demand Forecasting:** *'demand forecast of Paracetamol'* or *'fast selling medicines'*\n"
            "• **Medical Uses:** *'use of Paracetamol'*\n"
            "• **Side Effects & Disadvantages:** *'disadvantage of Paracetamol'*\n"
            "• **Dosage Guidelines:** *'dosage of Paracetamol'*\n"
            "• **Pregnancy Safety:** *'is Paracetamol safe during pregnancy'*\n"
            "• **Expiry Alerts:** *'what medicines are expiring soon'*\n"
            "• **Sales & Revenue:** *'Today\'s sales'* or *'Monthly revenue'*\n"
            "• You can also use the **Mic 🎤** to ask questions in English or Hindi!"
        )

    # 13. GENERAL SEARCH OR FALLBACK
    # 13a. Check if matches a customer
    matching_custs = _fuzzy_match_customers(raw_query, request=request)
    if matching_custs:
        return _format_customer_card(matching_custs, scoped_sales)

    # 13b. Check if matches a medicine
    matching_meds = _fuzzy_match_medicines(clean_query, request=request)
    if matching_meds:
        lines = []
        for m in matching_meds[:5]:
            batches = m.batches.filter(quantity__gt=0).order_by('expiry_date')
            b_info = ', '.join([f"`{b.batch_name}` (Qty: {b.quantity}, Exp: {b.expiry_date})" for b in batches]) if batches.exists() else "No active batches in stock"
            stock_badge = "🟢 In Stock" if m.total_quantity >= 50 else ("🟠 Low Stock" if m.total_quantity > 0 else "🔴 Out of Stock")
            lines.append(
                f"### 💊 {m.name}\n"
                f"• **Status:** {stock_badge} (**{m.total_quantity} units** in stock)\n"
                f"• **Selling Price:** **Rs.{m.price}**\n"
                f"• **Category:** {m.category.name if m.category else 'General'}\n"
                f"• **Supplier:** {m.supplier.name if m.supplier else 'N/A'}\n"
                f"• **Active Batches:** {b_info}"
            )
        return "\n\n".join(lines)

    # 13c. Check if matches a supplier
    matching_sups = _fuzzy_match_suppliers(clean_query, request=request)
    if matching_sups:
        lines = [f"• **{s.name}** (Phone: {s.contact_number or 'N/A'}, City: {s.location or 'N/A'}, Email: {s.email or 'N/A'})" for s in matching_sups[:5]]
        return f"🏢 **Supplier Details:**\n\n" + "\n".join(lines)

    return f"I could not find matching records for **'{raw_query}'**. Try asking about medicine stock (e.g. *'stock of Paracetamol'*), customer info (*'customer Neeraj Sharma'*), pricing, medical info (*'use of Amoxicillin'*), or expiry alerts (*'expiring soon'*)."


# ─────────────────────────────────────────────────────────────────────────────
@assistant_or_above
@require_POST
def chatbot_api(request):
    """
    Clean, robust MediBot API endpoint with AI intent classification,
    targeted context grounding, and clinical/inventory routing.
    """
    try:
        body = json.loads(request.body) if request.body else {}
        raw_msg = body.get('message', '').strip() if isinstance(body, dict) else ''
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'reply': 'Invalid request format.', 'status': 'error'}, status=400)

    if not raw_msg:
        return JsonResponse({'reply': 'Please type a question or speak into the microphone.', 'status': 'error'}, status=400)

    session = getattr(request, 'session', None)
    intent_info = _classify_query_intent(raw_msg, session=session)

    # 1. Call AI model (with intent framing and context)
    ai_reply = _call_gemini_api(raw_msg, intent_info=intent_info, request=request)

    # 2. Fallback to local rule/medical engine if AI call is unavailable
    final_reply_text = ai_reply if ai_reply else _medibot_local_engine(raw_msg, request=request)
    formatted_html = _format_markdown_to_html(final_reply_text)

    # Context entity persistence in session
    if session is not None:
        if intent_info.get('med_key'):
            session['medibot_last_med_key'] = intent_info['med_key']
        else:
            matched_meds = _fuzzy_match_medicines(raw_msg, request=request)
            if matched_meds:
                session['medibot_last_med_key'] = matched_meds[0].name.lower()

        matched_custs = _fuzzy_match_customers(raw_msg, request=request)
        if matched_custs:
            session['medibot_last_cust_name'] = matched_custs[0].name
        session.modified = True

    return JsonResponse({
        'reply': formatted_html,
        'raw_text': final_reply_text,
        'status': 'success'
    })




# SMART SALES ANALYTICS
# ─────────────────────────────────────────────────────────────────────────────

@assistant_or_above
def smart_analytics(request):
    """Dedicated Smart Sales Analytics dashboard"""
    return render(request, 'analytics/smart_analytics.html')


@assistant_or_above
def smart_analytics_data(request):
    """
    JSON API that powers the Smart Analytics dashboard.
    Returns KPIs, hourly heatmap, category revenue, payment mix, top customers.
    Query params: filter=today|week|month|year|custom  [&start_date=YYYY-MM-DD&end_date=YYYY-MM-DD]
    """
    from django.db.models import F, ExpressionWrapper, DecimalField, FloatField
    from django.db.models.functions import ExtractHour

    filter_type = request.GET.get('filter', 'month')
    start_date  = request.GET.get('start_date')
    end_date    = request.GET.get('end_date')
    today       = timezone.now().date()

    # ── Date range ────────────────────────────────────────────────────────────
    if filter_type == 'today':
        start = today
        end   = today
    elif filter_type == 'week':
        start = today - timedelta(days=7)
        end   = today
    elif filter_type == 'year':
        start = today - timedelta(days=365)
        end   = today
    elif filter_type == 'custom' and start_date and end_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            end   = datetime.strptime(end_date,   '%Y-%m-%d').date()
        except ValueError:
            start = today - timedelta(days=30)
            end   = today
    else:  # month (default)
        start = today - timedelta(days=30)
        end   = today

    # ── Completed sales in range, scoped to the active session owner ──────────────────────────────────────────────
    completed_sales = owner_scope_sales(
        request,
        Sale.objects.filter(
            date__date__gte=start, date__date__lte=end, status='Completed'
        )
    )
    all_sales = owner_scope_sales(
        request,
        Sale.objects.filter(date__date__gte=start, date__date__lte=end)
    )
    completed_items = SaleItem.objects.filter(
        sale__in=completed_sales,
    )

    # ── KPIs ──────────────────────────────────────────────────────────────────
    total_revenue = float(completed_sales.aggregate(r=Sum('total_price'))['r'] or 0)

    # Profit = sum(sell_price * qty) - sum(cost_price * qty) - discounts
    revenue_gross = float(
        completed_items.aggregate(
            r=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField()))
        )['r'] or 0
    )
    total_cost = float(
        completed_items.aggregate(
            c=Sum(ExpressionWrapper(F('cost_price') * F('quantity'), output_field=DecimalField()))
        )['c'] or 0
    )
    total_discount = float(completed_sales.aggregate(d=Sum('discount'))['d'] or 0)
    net_revenue = revenue_gross - total_discount
    total_profit = net_revenue - total_cost

    sale_count      = completed_sales.count()
    refund_count    = all_sales.filter(status='Refunded').count()
    total_count     = all_sales.count()
    refund_rate     = round((refund_count / total_count * 100) if total_count else 0, 1)
    avg_order_value = round(total_revenue / sale_count if sale_count else 0, 2)

    # Top medicine by units sold
    top_med_qs = (
        completed_items
        .values('medicine__name')
        .annotate(units=Sum('quantity'))
        .order_by('-units')
        .first()
    )
    top_medicine     = top_med_qs['medicine__name'] if top_med_qs else '—'
    top_medicine_qty = top_med_qs['units'] if top_med_qs else 0

    # ── Hourly heatmap ────────────────────────────────────────────────────────
    hourly_qs = (
        completed_sales
        .annotate(hour=ExtractHour('date'))
        .values('hour')
        .annotate(revenue=Sum('total_price'), count=Count('id'))
        .order_by('hour')
    )
    hours_map = {h: {'revenue': 0.0, 'count': 0} for h in range(24)}
    for row in hourly_qs:
        h = row['hour']
        hours_map[h] = {'revenue': float(row['revenue'] or 0), 'count': row['count']}
    hourly_data = {
        'hours':   list(range(24)),
        'revenue': [hours_map[h]['revenue'] for h in range(24)],
        'counts':  [hours_map[h]['count']   for h in range(24)],
    }

    # Peak hour
    peak_hour = max(range(24), key=lambda h: hours_map[h]['revenue']) if sale_count else 0
    def fmt_hour(h):
        suffix = 'am' if h < 12 else 'pm'
        display = h if h <= 12 else h - 12
        display = 12 if display == 0 else display
        return f"{display}{suffix}"

    # ── Revenue by category ───────────────────────────────────────────────────
    category_qs = (
        completed_items
        .values('medicine__category__name')
        .annotate(
            revenue=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField()))
        )
        .order_by('-revenue')
    )
    cat_labels  = [r['medicine__category__name'] or 'Uncategorised' for r in category_qs]
    cat_revenue = [float(r['revenue'] or 0) for r in category_qs]

    # ── Payment method mix ───────────────────────────────────────────────────
    payment_qs = (
        completed_sales
        .values('payment_method')
        .annotate(total=Sum('total_price'), count=Count('id'))
        .order_by('-total')
    )
    pay_labels  = [r['payment_method'] for r in payment_qs]
    pay_amounts = [float(r['total'] or 0) for r in payment_qs]
    pay_counts  = [r['count'] for r in payment_qs]

    # ── Revenue & Profit over time (daily) ────────────────────────────────────
    from django.db.models.functions import TruncDate as TruncD
    daily_revenue_qs = (
        completed_sales
        .annotate(d=TruncD('date'))
        .values('d')
        .annotate(rev=Sum('total_price'))
        .order_by('d')
    )
    daily_profit_qs = (
        completed_items
        .annotate(d=TruncD('sale__date'))
        .values('d')
        .annotate(
            rev=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())),
            cost=Sum(ExpressionWrapper(F('cost_price') * F('quantity'), output_field=DecimalField()))
        )
        .order_by('d')
    )
    discount_per_day = {
        row['d'].strftime('%Y-%m-%d'): float(row['disc'] or 0)
        for row in completed_sales
            .annotate(d=TruncD('date'))
            .values('d')
            .annotate(disc=Sum('discount'))
    }
    profit_map = {
        row['d'].strftime('%Y-%m-%d'): (float(row['rev'] or 0), float(row['cost'] or 0))
        for row in daily_profit_qs
    }
    trend_labels  = [row['d'].strftime('%Y-%m-%d') for row in daily_revenue_qs]
    trend_revenue = [float(row['rev'] or 0) for row in daily_revenue_qs]
    trend_profit  = []
    for lbl in trend_labels:
        rev_i, cost_i = profit_map.get(lbl, (0.0, 0.0))
        disc_i = discount_per_day.get(lbl, 0.0)
        trend_profit.append(round(rev_i - disc_i - cost_i, 2))

    # ── Top 5 customers ───────────────────────────────────────────────────────
    top_customers_qs = (
        completed_sales
        .filter(customer__isnull=False)
        .values('customer__name', 'customer__id')
        .annotate(spend=Sum('total_price'), orders=Count('id'))
        .order_by('-spend')[:5]
    )
    top_customers = [
        {
            'name':   r['customer__name'],
            'spend':  float(r['spend'] or 0),
            'orders': r['orders'],
        }
        for r in top_customers_qs
    ]

    # ── Smart insights ────────────────────────────────────────────────────────
    insights = []
    if sale_count == 0:
        insights.append('📭 No completed sales in this period. Try expanding the date range.')
    else:
        if total_profit > 0:
            margin = round(total_profit / net_revenue * 100, 1) if net_revenue else 0
            insights.append(f'💰 Profit margin is <strong>{margin}%</strong> — net profit of <strong>₹{total_profit:,.2f}</strong>.')
        elif total_profit < 0:
            insights.append(f'⚠️ Operating at a <strong>loss of ₹{abs(total_profit):,.2f}</strong> in this period.')

        if top_med_qs:
            share = round(top_medicine_qty / (completed_items.aggregate(t=Sum('quantity'))['t'] or 1) * 100, 1)
            insights.append(f'🏆 <strong>{top_medicine}</strong> is your top seller — {share}% of all units moved.')

        if peak_hour is not None and sale_count:
            peak_rev = hours_map[peak_hour]['revenue']
            insights.append(f'⏰ Peak sales hour is <strong>{fmt_hour(peak_hour)}</strong> with ₹{peak_rev:,.2f} in revenue.')

        if refund_rate > 10:
            insights.append(f'🔴 Refund rate is high at <strong>{refund_rate}%</strong> — consider investigating.')
        elif refund_rate > 0:
            insights.append(f'✅ Refund rate is healthy at <strong>{refund_rate}%</strong>.')

        if pay_labels:
            insights.append(f'💳 Most popular payment method: <strong>{pay_labels[0]}</strong> ({pay_amounts[0]:,.2f} ₹).')

        if cat_labels:
            insights.append(f'📦 Top category by revenue: <strong>{cat_labels[0]}</strong> (₹{cat_revenue[0]:,.2f}).')

    return JsonResponse({
        'kpi': {
            'total_revenue':    round(total_revenue, 2),
            'total_profit':     round(total_profit, 2),
            'avg_order_value':  avg_order_value,
            'sale_count':       sale_count,
            'refund_rate':      refund_rate,
            'top_medicine':     top_medicine,
            'top_medicine_qty': int(top_medicine_qty),
        },
        'trend': {
            'labels':  trend_labels,
            'revenue': trend_revenue,
            'profit':  trend_profit,
        },
        'hourly':     hourly_data,
        'categories': {'labels': cat_labels,  'data': cat_revenue},
        'payments':   {'labels': pay_labels,  'amounts': pay_amounts, 'counts': pay_counts},
        'top_customers': top_customers,
        'insights':   insights,
        'meta': {
            'start': start.strftime('%d %b %Y'),
            'end':   end.strftime('%d %b %Y'),
            'filter': filter_type,
        },
    })


@assistant_or_above
def demand_forecasting_view(request):
    """
    Renders the AI Demand Forecasting Dashboard.
    Displays:
      - 30-day predicted demand for all catalog medicines
      - Critical Stockout alerts & Reorder recommendations
      - Interactive Chart.js comparing past actual sales vs forecasted future demand
      - Dynamic dropdown to switch medicine forecast on the fly
    """
    all_forecasts = get_all_medicine_forecasts(days_ahead=30)
    medicines = Medicine.objects.order_by('name')

    selected_med_id = request.GET.get('med_id')
    if selected_med_id:
        try:
            selected_med_id = int(selected_med_id)
        except ValueError:
            selected_med_id = None

    if not selected_med_id and all_forecasts:
        selected_med_id = all_forecasts[0]['medicine_id']

    selected_forecast = None
    if selected_med_id:
        selected_forecast = forecast_demand(selected_med_id, days_ahead=30)

    # Executive Summary KPIs
    total_projected_units = round(sum(f['total_30_day_demand'] for f in all_forecasts), 1)
    critical_risk_count = sum(1 for f in all_forecasts if f['risk_level'] in ('CRITICAL', 'OUT_OF_STOCK'))
    high_risk_count = sum(1 for f in all_forecasts if f['risk_level'] == 'HIGH')
    moderate_risk_count = sum(1 for f in all_forecasts if f['risk_level'] == 'MODERATE')
    safe_count = sum(1 for f in all_forecasts if f['risk_level'] == 'ADEQUATE')
    total_reorder_units = sum(f['recommended_reorder_qty'] for f in all_forecasts)

    context = {
        'all_forecasts': all_forecasts,
        'medicines': medicines,
        'selected_med_id': selected_med_id,
        'selected_forecast': selected_forecast,
        'selected_forecast_json': json.dumps(selected_forecast),
        'total_projected_units': total_projected_units,
        'critical_risk_count': critical_risk_count,
        'high_risk_count': high_risk_count,
        'moderate_risk_count': moderate_risk_count,
        'safe_count': safe_count,
        'total_reorder_units': total_reorder_units,
        'catalog_count': len(all_forecasts),
    }
    return render(request, 'forecast/forecast_dashboard.html', context)


@assistant_or_above
def api_medicine_forecast(request, medicine_id):
    """
    JSON API for dynamic AJAX medicine switching on the forecasting chart.
    """
    data = forecast_demand(medicine_id, days_ahead=30)
    if not data:
        return JsonResponse({'error': 'Medicine not found'}, status=404)
    return JsonResponse(data)


@assistant_or_above
def trigger_critical_stock_alert_now(request):
    """
    Manual trigger view for immediate critical stock audit and email dispatch.
    """
    if request.method not in ('POST', 'GET'):
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    recipient = getattr(settings, 'EMAIL_HOST_USER', 'sharmaneeraj3415@gmail.com')
    # force=True so manual click always sends summary of at-risk / watchlist items
    result = send_forecast_critical_stock_email(recipient_email=recipient, force=True)
    return JsonResponse(result)


# =========================================================================
# 📅 FRONTEND EXPIRY REMINDER MODULE
# =========================================================================

def _get_expiry_reminder_queryset(request):
    """Return owner-scoped ExpiryReminderLog records respecting multi-tenant boundaries."""
    qs = ExpiryReminderLog.objects.all()
    if is_admin_role(request.user):
        return qs

    # For non-admin pharmacists, filter to customers/sales belonging to this pharmacist
    user_cust_emails = set(
        Customer.objects.filter(created_by=request.user)
        .exclude(email='')
        .values_list('email', flat=True)
    )
    user_sale_emails = set(
        Sale.objects.filter(created_by=request.user, customer__isnull=False)
        .exclude(customer__email='')
        .values_list('customer__email', flat=True)
    )
    all_scoped_emails = [e.strip().lower() for e in user_cust_emails.union(user_sale_emails) if e]

    if all_scoped_emails:
        return qs.filter(customer_email__in=all_scoped_emails)
    return qs


def _dispatch_consolidated_expiry_email(customer_name, customer_email, logs_list, conn=None):
    """
    Sends ONE consolidated branded email to the customer containing a clean table of all selected medicines.
    Ultra-fast: 1 single SMTP transmission regardless of how many medicines are selected.
    """
    email = (customer_email or '').strip()
    if not email or not logs_list:
        return False, "No email or medicines provided"

    customer_name = customer_name or 'Valued Customer'
    now_str = timezone.now().strftime("%d-%b-%Y %I:%M %p")
    today = timezone.now().date()
    ref_id = f"EXP-{logs_list[0].id}"

    if len(logs_list) == 1:
        subject = f"Your Medicine is Expiring Soon - {logs_list[0].medicine_name}"
    else:
        subject = f"Important Alert: {len(logs_list)} of your Medicines are Expiring Soon"

    # Build medicine rows for HTML & plain text
    rows_html = ""
    rows_plain = ""
    for log in logs_list:
        exp_str = log.expiry_date.strftime('%d-%b-%Y') if log.expiry_date else 'Soon'
        days_left = (log.expiry_date - today).days if log.expiry_date else None
        
        if days_left is not None:
            if days_left < 0:
                tag = f"Expired ({abs(days_left)}d ago)"
                color = "#b91c1c"
            elif days_left <= 7:
                tag = f"{days_left}d left (Critical)"
                color = "#dc2626"
            elif days_left <= 30:
                tag = f"{days_left}d left (Soon)"
                color = "#d97706"
            else:
                tag = f"{days_left}d left (Safe)"
                color = "#059669"
        else:
            tag = "No Date"
            color = "#64748b"

        rows_plain += f"- {log.medicine_name} | Expiry: {exp_str} ({tag})\n"
        rows_html += f"""
        <tr style="border-bottom: 1px solid #e2e8f0;">
          <td style="padding: 10px 12px; font-size: 13px; font-weight: 700; color: #0f172a;">💊 {log.medicine_name}</td>
          <td style="padding: 10px 12px; font-size: 13px; font-weight: 700; color: #334155; font-family: monospace;">{exp_str}</td>
          <td style="padding: 10px 12px; font-size: 12px; font-weight: 700; color: {color};">{tag}</td>
        </tr>
        """

    plain_body = (
        f"Dear {customer_name},\n\n"
        f"This is an important reminder from PharmaCare Pharmacy regarding your medicine(s) reaching expiry:\n\n"
        f"{rows_plain}\n"
        f"Please replace or restock your medicine in time to avoid any health risk or inconvenience.\n\n"
        f"Thank you,\n"
        f"PharmaCare Healthcare Team\n"
        f"Ref: {ref_id} | Dispatched: {now_str}"
    )

    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
      <meta charset="utf-8">
      <meta name="viewport" content="width=device-width, initial-scale=1.0">
    </head>
    <body style="margin:0; padding:16px; background-color:#f8fafc; font-family:'Segoe UI', Roboto, Helvetica, Arial, sans-serif;">
      <table width="100%" border="0" cellspacing="0" cellpadding="0" style="max-width:620px; margin:0 auto; background-color:#ffffff; border-radius:12px; overflow:hidden; border:1px solid #e2e8f0; box-shadow:0 4px 6px -1px rgba(0,0,0,0.08);">
        <tr>
          <td style="background:linear-gradient(135deg, #0891b2 0%, #0e7490 100%); padding:24px 28px; text-align:center; color:#ffffff;">
            <h1 style="margin:0; font-size:22px; font-weight:700; color:#ffffff;">PharmaCare Pharmacy</h1>
            <p style="margin:6px 0 0 0; font-size:13px; color:#e0f2fe;">Medicine Expiry Notification Alert ({len(logs_list)} Item{'s' if len(logs_list)>1 else ''})</p>
          </td>
        </tr>
        <tr>
          <td style="padding:28px 28px 20px 28px; color:#334155; line-height:1.6;">
            <p style="font-size:16px; font-weight:600; color:#0f172a; margin-top:0;">Dear {customer_name},</p>
            <p style="font-size:14px; color:#475569; margin:12px 0 16px 0;">
              This is an important reminder regarding your medicine(s) reaching expiry:
            </p>
            
            <table width="100%" border="0" cellspacing="0" cellpadding="0" style="background:#f8fafc; border-radius:8px; border:1px solid #cbd5e1; margin:16px 0; overflow:hidden;">
              <thead>
                <tr style="background:#e2e8f0; text-align:left;">
                  <th style="padding:8px 12px; font-size:11px; font-weight:700; color:#475569; text-transform:uppercase;">Medicine</th>
                  <th style="padding:8px 12px; font-size:11px; font-weight:700; color:#475569; text-transform:uppercase;">Expiry Date</th>
                  <th style="padding:8px 12px; font-size:11px; font-weight:700; color:#475569; text-transform:uppercase;">Status</th>
                </tr>
              </thead>
              <tbody>
                {rows_html}
              </tbody>
            </table>

            <p style="font-size:14px; color:#475569; margin:18px 0;">
              Please consult your pharmacist or doctor to replace or restock in time to avoid health risks.
            </p>
            
            <p style="margin:24px 0 0 0; font-size:14px; color:#334155;">
              Thank you,<br>
              <strong style="color:#0891b2;">PharmaCare Healthcare Team</strong>
            </p>
          </td>
        </tr>
        <tr>
          <td style="background:#f1f5f9; padding:14px 28px; text-align:center; font-size:11px; color:#94a3b8; border-top:1px solid #e2e8f0;">
            PharmaCare Management System • Reference #{ref_id} • Sent at {now_str}
          </td>
        </tr>
      </table>
    </body>
    </html>
    """

    try:
        from MediApp.utils.email_utils import send_universal_mail
        success, err = send_universal_mail(
            subject=subject,
            plain_body=plain_body,
            html_body=html_body,
            to_email=email,
            conn=conn
        )
        if success:
            for log in logs_list:
                log.reminder_sent = True
                log.sent_at = timezone.now()
                log.message = plain_body
                log.save(update_fields=['reminder_sent', 'sent_at', 'message'])
            return True, None
        else:
            for log in logs_list:
                log.message = f"Send Failed: {err}"
                log.save(update_fields=['message'])
            return False, str(err)
    except Exception as exc:
        for log in logs_list:
            log.message = f"Send Failed: {exc}"
            log.save(update_fields=['message'])
        return False, str(exc)


def _dispatch_expiry_email_for_log(log, conn=None):
    """
    Dispatches reminder email for a single log.
    """
    return _dispatch_consolidated_expiry_email(log.customer_name, log.customer_email, [log], conn=conn)


@assistant_or_above
def expiry_reminder_list(request):
    """
    Frontend Expiry Reminder dashboard listing for pharmacists.
    Shows medicines expiring soon/past expiry with bulk-select and email trigger.
    """
    # Auto-synchronize alerts on page load if needed
    try:
        from MediApp.admin import sync_customer_expiry_logs
        sync_customer_expiry_logs()
    except Exception:
        pass

    today = timezone.now().date()
    qs = _get_expiry_reminder_queryset(request)

    # Filters
    search_query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', 'all')
    timeline_filter = request.GET.get('timeline', 'all')

    if search_query:
        qs = qs.filter(
            Q(customer_name__icontains=search_query) |
            Q(customer_email__icontains=search_query) |
            Q(medicine_name__icontains=search_query)
        )

    if status_filter == 'pending':
        qs = qs.filter(reminder_sent=False)
    elif status_filter == 'sent':
        qs = qs.filter(reminder_sent=True)

    if timeline_filter == 'critical':
        # <= 7 days left (or expired)
        qs = qs.filter(expiry_date__lte=today + timedelta(days=7))
    elif timeline_filter == 'soon':
        # <= 30 days left
        qs = qs.filter(expiry_date__gte=today, expiry_date__lte=today + timedelta(days=30))
    elif timeline_filter == 'expired':
        qs = qs.filter(expiry_date__lt=today)
    elif timeline_filter == 'safe':
        qs = qs.filter(expiry_date__gt=today + timedelta(days=30))

    # KPI Metrics across entire scoped set
    base_qs = _get_expiry_reminder_queryset(request)
    total_alerts_count = base_qs.count()
    pending_alerts_count = base_qs.filter(reminder_sent=False).count()
    sent_alerts_count = base_qs.filter(reminder_sent=True).count()
    critical_alerts_count = base_qs.filter(expiry_date__gte=today, expiry_date__lte=today + timedelta(days=7)).count()
    expired_alerts_count = base_qs.filter(expiry_date__lt=today).count()

    # Ordering: soonest expiring first
    qs = qs.order_by('expiry_date', 'customer_name')
    filtered_count = qs.count()

    # Pagination: 25 items per page
    page = request.GET.get('page', 1)
    paginator = Paginator(qs, 25)
    try:
        logs_page = paginator.page(page)
    except PageNotAnInteger:
        logs_page = paginator.page(1)
    except EmptyPage:
        logs_page = paginator.page(paginator.num_pages)

    # Fetch customer phone numbers for records on this page
    import urllib.parse
    import re
    cust_emails = [item.customer_email.strip().lower() for item in logs_page if item.customer_email]
    cust_names = [item.customer_name.strip() for item in logs_page if item.customer_name]

    phone_map = {}
    if cust_emails:
        for c in Customer.objects.filter(email__in=cust_emails).exclude(contact_number=''):
            phone_map[c.email.strip().lower()] = c.contact_number
    if cust_names:
        for c in Customer.objects.filter(name__in=cust_names).exclude(contact_number=''):
            if c.name.strip() not in phone_map:
                phone_map[c.name.strip()] = c.contact_number

    # Attach computed timeline metadata and WhatsApp links to each log item in the page
    for item in logs_page:
        email_clean = (item.customer_email or '').strip().lower()
        name_clean = (item.customer_name or '').strip()
        phone = phone_map.get(email_clean) or phone_map.get(name_clean) or ''
        item.customer_phone = phone

        # Format number for WhatsApp link
        clean_num = re.sub(r'[^0-9]', '', phone)
        if len(clean_num) == 10:
            clean_num = '91' + clean_num
        item.whatsapp_phone = clean_num

        if item.expiry_date:
            days_left = (item.expiry_date - today).days
            item.days_left = days_left
            if days_left < 0:
                item.badge_class = 'bg-rose-100 text-rose-800 border-rose-300'
                item.badge_text = f'Expired ({abs(days_left)}d ago)'
                item.badge_icon = '🔴'
            elif days_left <= 7:
                item.badge_class = 'bg-red-100 text-red-800 border-red-300'
                item.badge_text = f'{days_left}d left (Critical)'
                item.badge_icon = '⚠️'
            elif days_left <= 30:
                item.badge_class = 'bg-amber-100 text-amber-800 border-amber-300'
                item.badge_text = f'{days_left}d left (Expiring Soon)'
                item.badge_icon = '🟠'
            else:
                item.badge_class = 'bg-emerald-100 text-emerald-800 border-emerald-300'
                item.badge_text = f'{days_left}d left (Safe)'
                item.badge_icon = '🟢'
        else:
            item.days_left = None
            item.badge_class = 'bg-gray-100 text-gray-600 border-gray-300'
            item.badge_text = 'No Date'
            item.badge_icon = '⚪'

        exp_str = item.expiry_date.strftime('%d-%b-%Y') if item.expiry_date else 'Soon'
        cust_display_name = item.customer_name or 'Valued Customer'
        status_info = item.badge_text

        wa_msg = (
            f"🏥 *PharmaCare Pharmacy - Medicine Expiry Alert*\n\n"
            f"Dear *{cust_display_name}*,\n"
            f"This is an important reminder regarding your prescribed medicine:\n\n"
            f"💊 *Medicine:* {item.medicine_name}\n"
            f"📅 *Expiry Date:* {exp_str}\n"
            f"⚠️ *Status:* {status_info}\n\n"
            f"Please visit PharmaCare Pharmacy or consult your doctor to replace or restock your supply in time.\n\n"
            f"Stay Healthy,\n"
            f"*PharmaCare Healthcare Team*"
        )
        item.whatsapp_message = wa_msg
        item.whatsapp_url = f"https://wa.me/{clean_num}?text={urllib.parse.quote(wa_msg)}" if clean_num else ""

    context = {
        'logs': logs_page,
        'page_obj': logs_page,
        'paginator': paginator,
        'is_paginated': logs_page.has_other_pages(),
        'total_alerts_count': total_alerts_count,
        'pending_alerts_count': pending_alerts_count,
        'sent_alerts_count': sent_alerts_count,
        'critical_alerts_count': critical_alerts_count,
        'expired_alerts_count': expired_alerts_count,
        'filtered_count': filtered_count,
        'search_query': search_query,
        'status_filter': status_filter,
        'timeline_filter': timeline_filter,
    }
    return render(request, 'expiry/expiry_reminder_list.html', context)


@assistant_or_above
def send_expiry_reminders_bulk(request):
    """
    Bulk email trigger endpoint for selected customer expiry logs.
    """
    if request.method != 'POST':
        return redirect('expiry_reminder_list')

    selected_ids = request.POST.getlist('selected_logs')
    if not selected_ids:
        messages.warning(request, "⚠️ Please select at least one customer record to send reminders.")
        return redirect('expiry_reminder_list')

    qs = _get_expiry_reminder_queryset(request).filter(id__in=selected_ids)
    
    # Group selected medicines by customer email to send 1 consolidated email per customer
    from collections import defaultdict
    cust_groups = defaultdict(list)
    for log in qs:
        email = (log.customer_email or '').strip().lower()
        if not email:
            continue
        cust_groups[email].append(log)

    sent_count = 0
    skipped_count = len(qs) - sum(len(items) for items in cust_groups.values())
    failed_count = 0

    conn = None
    try:
        from MediApp.admin import get_email_connection
        conn = get_email_connection()
    except Exception as exc:
        conn = None

    last_error = None
    for email, items in cust_groups.items():
        cust_name = items[0].customer_name or 'Valued Customer'
        try:
            success, err = _dispatch_consolidated_expiry_email(cust_name, email, items, conn=conn)
            if success:
                sent_count += 1
            else:
                failed_count += 1
                last_error = err
        except Exception as exc:
            failed_count += 1
            last_error = str(exc)

    total_meds_sent = sum(len(items) for email, items in cust_groups.items()) if sent_count > 0 else 0
    if sent_count > 0 and failed_count == 0 and skipped_count == 0:
        messages.success(request, f"✅ Successfully sent consolidated expiry reminder email(s) to {sent_count} customer(s) ({total_meds_sent} medicine alerts)!")
    elif sent_count > 0:
        messages.info(request, f"📧 Reminder Email Results: {sent_count} customer(s) notified ({total_meds_sent} medicines), {skipped_count} skipped (no email), {failed_count} failed ({last_error}).")
    else:
        messages.error(request, f"❌ Failed to send emails ({failed_count} failed, {skipped_count} skipped without email on file. Reason: {last_error or 'Check SMTP settings'}).")

    # Retain filter query parameters in redirect if provided
    redirect_url = request.POST.get('redirect_url')
    if redirect_url:
        return redirect(redirect_url)
    return redirect('expiry_reminder_list')


@assistant_or_above
def send_single_expiry_reminder(request, log_id):
    """
    Send reminder email to a single customer record.
    """
    log = get_object_or_404(_get_expiry_reminder_queryset(request), id=log_id)
    success, err = _dispatch_expiry_email_for_log(log)
    if success:
        messages.success(request, f"✅ Expiry reminder email sent successfully to {log.customer_name} ({log.customer_email})!")
    else:
        messages.error(request, f"❌ Failed to send email to {log.customer_email}: {err}")
    return redirect('expiry_reminder_list')


@assistant_or_above
def sync_expiry_reminders_view(request):
    """
    Manually scans sales history to discover and sync all upcoming expiry alert records.
    """
    try:
        from MediApp.admin import sync_customer_expiry_logs
        new_count = sync_customer_expiry_logs()
        messages.success(request, f"✅ Synchronized customer expiry records! {new_count} new alert(s) discovered from sales history.")
    except Exception as exc:
        messages.error(request, f"❌ Synchronization error: {exc}")
    return redirect('expiry_reminder_list')


@assistant_or_above
def api_customer_medicines_for_reminder(request):
    """
    JSON endpoint for smart autocomplete & purchase history retrieval for the Expiry Reminder modal.
    Accepts customer_id or query string (name/phone/email).
    Returns customer metadata + unique list of medicines purchased by the customer with batch & expiry info.
    """
    customer_id = request.GET.get('customer_id')
    query = request.GET.get('q', '').strip()

    today = timezone.now().date()
    customers_qs = owner_scope_queryset(request, Customer.objects.all(), 'created_by')

    if customer_id:
        customer = customers_qs.filter(id=customer_id).first()
        if not customer:
            return JsonResponse({'success': False, 'error': 'Customer not found'}, status=404)

        # Get all medicines ever purchased by this customer with batch expiry dates
        sale_items = (
            SaleItem.objects.filter(sale__customer=customer, batch__isnull=False)
            .select_related('medicine', 'batch', 'sale')
            .order_by('-sale__date')
        )

        purchased_meds = []
        seen_keys = set()

        for item in sale_items:
            med_name = item.medicine.name if item.medicine else 'Unknown Medicine'
            batch_name = item.batch.batch_name if item.batch else 'General'
            exp_date = item.batch.expiry_date
            exp_str = exp_date.strftime('%Y-%m-%d') if exp_date else ''
            
            key = f"{med_name}|{exp_str}"
            if key in seen_keys:
                continue
            seen_keys.add(key)

            days_left = (exp_date - today).days if exp_date else None
            status_tag = 'Safe'
            status_color = 'emerald'
            if days_left is not None:
                if days_left < 0:
                    status_tag = f'Expired ({abs(days_left)}d ago)'
                    status_color = 'rose'
                elif days_left <= 7:
                    status_tag = f'{days_left}d left (Critical)'
                    status_color = 'red'
                elif days_left <= 30:
                    status_tag = f'{days_left}d left (Expiring Soon)'
                    status_color = 'amber'
                else:
                    status_tag = f'{days_left}d left'
                    status_color = 'emerald'

            purchased_meds.append({
                'medicine_name': med_name,
                'batch_name': batch_name,
                'expiry_date': exp_str,
                'expiry_display': exp_date.strftime('%d-%b-%Y') if exp_date else 'No Date',
                'days_left': days_left,
                'status_tag': status_tag,
                'status_color': status_color,
                'purchase_date': item.sale.date.strftime('%d-%b-%Y') if item.sale else '',
                'quantity': item.quantity,
            })

        return JsonResponse({
            'success': True,
            'customer': {
                'id': customer.id,
                'name': customer.name,
                'email': customer.email,
                'contact_number': customer.contact_number,
            },
            'purchased_medicines': purchased_meds,
        })

    elif query:
        # Search autocomplete for customer name, phone, email
        matches = customers_qs.filter(
            Q(name__icontains=query) |
            Q(contact_number__icontains=query) |
            Q(email__icontains=query)
        )[:8]

        results = [
            {
                'id': c.id,
                'name': c.name,
                'email': c.email,
                'contact_number': c.contact_number,
            }
            for c in matches
        ]
        return JsonResponse({'success': True, 'results': results})

    # Default: return list of all active catalog medicines in pharmacy
    medicines = owner_scope_queryset(request, Medicine.objects.filter(is_active=True).prefetch_related('batches'), 'created_by')[:50]
    med_list = []
    for m in medicines:
        nearest = m.nearest_active_batch
        med_list.append({
            'id': m.id,
            'name': m.name,
            'batch_name': nearest.batch_name if nearest else '',
            'expiry_date': nearest.expiry_date.strftime('%Y-%m-%d') if (nearest and nearest.expiry_date) else '',
        })
    return JsonResponse({'success': True, 'catalog_medicines': med_list})


@assistant_or_above
def create_and_send_customer_expiry_reminders(request):
    """
    Creates ExpiryReminderLog records for selected medicines and dispatches reminder email(s).
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST method required'}, status=405)

    try:
        data = json.loads(request.body)
    except Exception:
        data = request.POST

    customer_name = (data.get('customer_name') or 'Valued Customer').strip()
    customer_email = (data.get('customer_email') or '').strip().lower()
    send_email_now = data.get('send_email_now', True)
    if isinstance(send_email_now, str):
        send_email_now = send_email_now.lower() in ('true', '1', 'yes')

    medicines = data.get('medicines', [])
    if isinstance(medicines, str):
        try:
            medicines = json.loads(medicines)
        except Exception:
            medicines = []

    if not customer_email:
        return JsonResponse({'success': False, 'error': 'Customer email address is required to create reminder'}, status=400)

    if not medicines:
        return JsonResponse({'success': False, 'error': 'Please select at least one medicine for the reminder'}, status=400)

    created_logs = []
    for item in medicines:
        med_name = (item.get('medicine_name') or '').strip()
        exp_str = item.get('expiry_date')
        if not med_name:
            continue

        exp_date = None
        if exp_str:
            try:
                exp_date = datetime.strptime(exp_str[:10], '%Y-%m-%d').date()
            except Exception:
                exp_date = timezone.now().date()
        else:
            exp_date = timezone.now().date()

        log, _ = ExpiryReminderLog.objects.get_or_create(
            customer_email=customer_email,
            medicine_name=med_name,
            expiry_date=exp_date,
            defaults={
                'customer_name': customer_name,
                'reminder_sent': False,
            }
        )
        if log.customer_name != customer_name:
            log.customer_name = customer_name
            log.save(update_fields=['customer_name'])
        created_logs.append(log)

    sent_count = 0
    err_msg = None

    if send_email_now:
        success, err = _dispatch_consolidated_expiry_email(customer_name, customer_email, created_logs)
        if success:
            sent_count = len(created_logs)
        else:
            err_msg = err

    msg = f"✅ Reminder{'s' if len(created_logs)>1 else ''} saved! "
    if send_email_now:
        if sent_count > 0:
            msg += f"Sent 1 consolidated email to {customer_email} with all {sent_count} medicine alert{'s' if sent_count>1 else ''}."
        else:
            msg += f"Could not send email ({err_msg if err_msg else 'Check mail settings'})."

    return JsonResponse({
        'success': True,
        'message': msg,
        'created_count': len(created_logs),
        'sent_count': sent_count,
    })


